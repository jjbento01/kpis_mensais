import pandas as pd
import concurrent.futures
import threading
from yaml import load, FullLoader
import codecs
import os
from sqlalchemy import engine, create_engine
from datetime import datetime, timedelta
#import pandas as pd
import polars as pl
from openpyxl import Workbook
from openpyxl import worksheet
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import NamedStyle, Side
from icecream import ic

from lista_queries import *
from utilities import *
from formatos import *
from logins import *
from logins_tv import *
from novos_users_ac import *
from pandas.core.indexes.datetimes import date_range



with codecs.open(str(os.getenv("CONFS"))+"conf_dba\\configuracao.yaml", "r", "utf-8") as fich:
    config = load(fich, Loader=FullLoader)
    string_conf = config['con_str_0']

bd: str ='BD_ECARE'
engine: engine = create_engine(f'mssql+pyodbc:///?odbc_connect={ string_conf };DATABASE='+bd)
dados: dict[str, pl.DataFrame]={}

def download_site(query: list[str,str])->None:
    global engine, dados
    ic(query)
    dados[query[0]] = pl.read_database(query[1], engine)

def download_all_sites(lista_queries: list[list[str,str]])->None:
    with concurrent.futures.ThreadPoolExecutor(max_workers=7) as executor:
        executor.map(download_site, lista_queries)


def do_main()->tuple:
    global lista_queries_a_fazer, lista, lista_unica,dados
    download_all_sites(lista_queries_a_fazer)
    date_range = range_needed_of_date(datetime.today())
    [ultimo, _, primeiro_ultimo_mes, first_last_month, last_last_month] = calc_date_values()
    [[first_day_week, first_year, first_week],
     [second_day_week, second_year, second_week],
     [third_day_week, third_year, third_week],
     [fourth_day_week, fourth_year, fourth_week],
     [fifth_day_week, fifth_year, fifth_week]] = getdays_for_week(ultimo)
    ic(date_range[0].strftime("%Y-%m-%d"))
    ic(primeiro_ultimo_mes)
    ic(date_range[-1].strftime("%Y-%m-%d"))
    ic(first_last_month.strftime("%Y-%m-%d"))
    ic(last_last_month.strftime("%Y-%m-%d"))
    ic("=====================================")
    wb = Workbook()
    #carregar_format(wb)
    ws = wb.active
    ws.title="Report"
    ws.column_dimensions['A'].width=5
    ws.column_dimensions['B'].width=30
    tags = [
        'Área de Cliente',
        '',
        'Logins Web',
        '      Logins My Meo',
        '      Logins Moche',
        '      Logins Uzo',
        '      Logins My PT Empresas',
        '      Logins AC PT Empresas',
        '      Unique Logins',
        '      Average Logins per User']
    estilos = [cabecalho, cabecalho, total, normal, normal, normal, normal, normal, normal, normal]
    bloco_principal_logins(ws,
                           first_last_month,
                           last_last_month,
                           primeiro_ultimo_mes,
                           ultimo,
                           lista_unica,
                           date_range,
                           dados,
                           tags,
                           estilos)
    for coluna, (iweek, iday) in zip(range(7,13), ((first_week, first_day_week),
                                                   (second_week, second_day_week),
                                                   (third_week, third_day_week),
                                                   (fourth_week, fourth_day_week),
                                                   (fifth_week, fifth_day_week))):
        ws.cell(row=2, column=len(date_range)+coluna, value="W"+str(iweek)).style=totalinhadir
        ws.cell(
            row=3,
            column=len(date_range)+coluna,
            value=(iday - datetime(1899, 12, 30)).days
        ).style=normalshort
    medias: dict[str, float] = {item: 0 for item in lista}
    for linha, item in zip([7, 8, 9, 10, 11, 12], lista):
        for ipos, (year, week) , sty in zip(range(7, 13),
                                          [(first_year, first_week),
                                           (second_year, second_week),
                                           (third_year, third_week),
                                           (fourth_year, fourth_week),
                                           (fifth_year, fifth_week)],
                                          [nmgrds, nmgrds, nmgrds, nmgrds, nmgrds]):
            temp = get_data_year_week(dados, 'logins_week', year, week, item)
            ws.cell(row=linha, column=len(date_range)+ipos, value=temp).style=sty
            #import ipdb; ipdb.set_trace()
            medias[item]+=temp if not(year==fifth_year and week==fifth_week) else 0

    per_user_list: list[float] = []

    valores_totais: list[float] = []

    for coluna, (year, week) in zip(
        range(7, 12),
        [(first_year, first_week),
         (second_year, second_week),
         (third_year, third_week),
         (fourth_year, fourth_week),
         (fifth_year, fifth_week)]):
        temp = 0
        for item in lista_unica:
            temp += get_data_year_week(dados, 'logins_week', year, week, item)
        valores_totais.append(temp)
        per_user= temp/get_data_year_week(dados, 'logins_week', year, week, 'Qtd_Logins_Unique')
        per_user_list.append(per_user)
        ws.cell(row=6, column=len(date_range)+coluna, value=temp).style=totalinha
        ws.cell(row=13, column=len(date_range)+coluna, value=per_user).style=normalunder

    coluna: int = len(date_range)+12

    total_fifth: int = 0
    total_fourth: int = 0

    for linha, item, color in zip(
        range(len(lista_unica)),
        lista_unica,
        [
            normalgrayperc,
            normalgrayperc,
            normalgrayperc,
            normalgrayperc,
            normalgrayperc,
            normalgrayperc
        ]):
        temp_fifth: int = get_data_year_week(dados, 'logins_week', fifth_year, fifth_week, item)
        temp_fourth: int = get_data_year_week(dados, 'logins_week', fourth_year, fourth_week, item)
        rto: float = 0 if temp_fourth == 0 else temp_fifth / temp_fourth - 1
        ws.cell(row=7+linha, column=coluna, value=rto).style=color
        total_fifth += temp_fifth
        total_fourth += temp_fourth

    resultado: float = get_data_year_week(dados,
        'logins_week',
        fifth_year,
        fifth_week,
        'Qtd_Logins_Unique'
    )/get_data_year_week(dados,
        'logins_week',
        fourth_year,
        fourth_week,
        'Qtd_Logins_Unique') - 1

    ic(resultado)

    ws.cell(row=7+len(lista_unica), column=coluna, value=(
        get_data_year_week(dados,
            'logins_week',
            fifth_year,
            fifth_week,
            'Qtd_Logins_Unique'
        )/get_data_year_week(dados,
            'logins_week',
            fourth_year,
            fourth_week,
            'Qtd_Logins_Unique'
        ) - 1)).style=normalgrayperc

    ws.cell(row=6, column=coluna, value=total_fifth/total_fourth-1).style=tlinhagrayperc
    ws.cell(row=13, column=coluna, value=per_user_list[-1]/per_user_list[-2]-1).style=normalgrayunderperc

    for i, primeiralinha, segundalinha in zip(range(3), ("W"+str(first_week)+" vs.", "", ""), ("1w", "Avg.4w", "Avg.YTD")):
        ws.cell(row=2, column=coluna+i, value=primeiralinha).style=tlinhagrayperc
        ws.cell(row=3, column=coluna+i, value=segundalinha).style=normalgrayunder

    valores_inter = [
        dados['logins_week'].filter(
            (
                dados['logins_week']['Year']==fifth_year
            )&(
                dados['logins_week']['Week']==fifth_week
            )
        )[item]/(medias[item]/4)-1 for item in lista
    ]
    total_val = valores_totais[-1]/(sum([medias[key] for key in medias.keys()][:-1])/4)-1
    ws.cell(row=6, column=coluna+1, value=total_val).style=tlinhagrayperc
    for i, valor in enumerate(valores_inter): ws.cell(row=7+i, column=coluna+1, value=list(valor)[0]).style=normalgrayperc
    total_fim = per_user_list[-1]/(sum(per_user_list[:-1])/4)-1
    ws.cell(row=13, column=coluna+1, value=total_fim).style=normalgrayunderperc
    coluna_ytd_semanal_web_logins('logins_week', "Qtd_Logins_Unique", fifth_year, fifth_week, coluna, dados, ws, lista_unica, lista)
    letra_no_fim = ws.cell(row=5, column=(len(date_range)-5)).column_letter
    ws.column_dimensions.group('C', letra_no_fim, hidden=True)
    ws.column_dimensions[ws.cell(row=2, column=coluna+3).column_letter].width=3
    coluna += 4
    coluna_inicial = coluna
    lista_logins_users = colunas_mensais_web_logins(dados, ws, coluna)
    coluna += 12
    letra_inicial = ws.cell(row=5, column=coluna_inicial).column_letter
    letra_final = ws.cell(row=5, column=coluna_inicial+7).column_letter
    ws.column_dimensions.group(letra_inicial, letra_final, hidden=True)
    coluna_mensal_um_mes(dados, ws, coluna, lista)
    coluna+=1
    coluna_mensal_4_meses(dados, ws, coluna, lista, lista_logins_users)
    coluna+=1
    coluna_ytd_mensal_logins(dados, ws, coluna, lista)
    coluna+=1
    ws.column_dimensions[ws.cell(row=2, column=coluna).column_letter].width=3
    coluna+=1
    coluna_vazia(ws,coluna)
    row_actual = 14
    coluna = 2
    cabecalho_tv_logins(ws, coluna, row_actual)
    row_actual = 15
    coluna += 1
    valores_diarios_tv_logins(dados, ws, coluna, row_actual)
    coluna += dados['logins_day_tv'].shape[0]
    primeiro_dia_mes = (
        datetime(datetime.now().year, datetime.now().month, 1) - timedelta(days=1)
    ).replace(day=1)
    ultimo_dia_mes = datetime(datetime.now().year, datetime.now().month, 1) - timedelta(days=1)
    lista_fla = total_mes_analise(dados, ws, coluna, row_actual, primeiro_dia_mes, ultimo_dia_mes)
    coluna += 1
    primeiro_dia_anterior = (primeiro_dia_mes - timedelta(days=1)).replace(day=1)
    ultimo_dia_anterior = edate(ultimo_dia_mes)
    lista_fll = total_mes_analise(
        dados,
        ws,
        coluna,
        row_actual,
        primeiro_dia_anterior,
        ultimo_dia_anterior)
    coluna += 1
    total_mes_analise_var(ws, coluna, row_actual, lista_fla, lista_fll)
    coluna +=3
    calculate_week_columns(
        dados, 'logins_week_tv', 'Qtd_Visits_Unique',
        ws, coluna, date_range, first_year, first_week,
        second_year, second_week, third_year, third_week,
        fourth_year, fourth_week, fifth_year, fifth_week,
        ['Qtd_Visits', 'Qtd_Visits_Intentional'], 15, 17)
    coluna += 4
    get_w48vs1w(dados, "logins_week_tv",
        ['Qtd_Visits', 'Qtd_Visits_Intentional', 'Qtd_Visits_Unique'], ws,
        fifth_year, fifth_week, fourth_year, fourth_week, 15, coluna)
    coluna += 1
    get_w48vs4w(dados, 'logins_week_tv',
                ['Qtd_Visits', 'Qtd_Visits_Intentional', 'Qtd_Visits_Unique'],
                ws, fifth_year, fifth_week, fourth_year, fourth_week,
                third_year, third_week, second_year, second_week,
                first_year, first_week, 15, coluna)
    coluna += 1
    tv_ytd_semanal_web_logins(
        'logins_week_tv',
        fifth_year,
        fifth_week,
        coluna,
        dados,
        ws,
        15,
        ['Qtd_Visits', 'Qtd_Visits_Intentional', 'Qtd_Visits_Unique'])
    coluna += 2
    ret: list[float] = colunas_mensais_tv_logins(
        dados,
        'logins_month_tv',
        ['Qtd_Visits', 'Qtd_Visits_Intentional', 'Qtd_Visits_Unique'],
        ws,
        15,
        coluna)
    coluna += 12
    tv_make_last_month(
        dados,
        ws,
        coluna,
        ['Qtd_Visits', 'Qtd_Visits_Intentional', 'Qtd_Visits_Unique', ''],
        ret)
    coluna += 1
    tv_make_last_four_month(
        dados,
        ws,
        coluna,
        ['Qtd_Visits', 'Qtd_Visits_Intentional', 'Qtd_Visits_Unique'],
        ret)
    coluna += 1
    coluna_ytd_mensal_tv_logins(
        'logins_month_tv',
        fifth_year,
        ultimo.month,
        coluna,
        dados,
        ws,
        ['Qtd_Visits', 'Qtd_Visits_Intentional', 'Qtd_Visits_Unique'])
    coluna += 2
    tv_coluna_vazia(ws, 15, coluna, ['Qtd_Visits', 'Qtd_Visits_Intentional', 'Qtd_Visits_Unique', ''])
    coluna = 2
    novos_users_ac_cabecalhos(ws, coluna, 19)
    coluna += 1
    valores_diarios_new_users("new_users_day", dados, ws, coluna, 20, normal, nmgrds, nmgrdsund)
    coluna += valores_diarios_users(dados, ws, coluna, 21)
    month_total(dados, ws, coluna, 21)
    #coluna += 4
    #month_col = coluna
    # weeks_in_year(
    #     [
    #         [first_year, first_week],
    #         [second_year, second_week],
    #         [third_year, third_week],
    #         [fourth_year, fourth_week],
    #         [fifth_year, fifth_week]
    #     ],
    #     21,
    #     coluna,
    #     ws,
    #     dados,
    #     "users_ac_week",
    #     "new_users_week"
    # )
    # coluna += 1
    # week_summary_1([[fifth_year, fifth_week], [fourth_year, fourth_week]], 20, coluna, ws, dados, "users_ac_week", "new_users_week")
    # month_col += 9
    ic(ultimo)
    primeiro_ano = (ultimo - pd.DateOffset(years=1) ) + timedelta(days=1)
    ic(primeiro_ano)
    #months_in_year(fifth_year, primeiro_ano.month, first_year, 21, month_col, dados, ws)
    wb.save("sample.xlsx")
    engine.dispose()
    return (dados, valores_inter, medias)

if __name__ == '__main__':
    dados, valores_inter, medias = do_main()
