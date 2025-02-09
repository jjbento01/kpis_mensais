import polars as pl
from formatos import *
from utilities import *
from openpyxl import worksheet
from datetime import datetime, timedelta

def novos_users_ac_cabecalhos(ws, coluna, row_actual):
    tags = [
    'Logins',
    '      New First Logins',
    '      ACC Total*',
    '            ACC Meo',
    '            ACC Moche*****',
    '            ACC UZO*****',
    '            ACC Altice Empresas****',
    '      ACC ∩ ACE',
    '      ACE Total**',
    '            ACE só Empresariais',
    '            ACE Altice Empresas',
    '      Altice Empresas Total']
    estilos = [total, normal, normal, normal, normal, normal, normal, normal, normal, normal, normal, normal]
    for i, tag, sty in zip(range(0, 13), tags[:-1], estilos):
        ws.cell(row=row_actual+i, column=coluna,  value=tag).style=sty
    ws.cell(row=row_actual+i+1, column=coluna, value=tags[-1]).style=normalunder

def valores_diarios_new_users(tabela: str, dados: dict, ws: worksheet, coluna: int, row_actual: int)->None:
    primeiro = first_day_a_year_ago(datetime.today().date())
    ultimo = last_day_of_last_month(datetime.today()).date()
    pdd = dados[tabela].filter((pl.col('Date')>=primeiro)&(pl.col('Date')<=ultimo)).sort(by="Date")
    for i, row in enumerate(pdd.iter_rows()):
        ws.cell(row=row_actual, column=coluna+i, value=row[1]).style=normal

def valores_diarios_users(dados: dict, ws: worksheet, coluna: int, row_actual: int)->int:
    primeiro = first_day_a_year_ago(datetime.today().date())
    ultimo = last_day_of_last_month(datetime.today()).date()
    pcd = dados['new_users_day'].sort(by='Date')
    pdd = dados['users_ac_day'].sort(by="Date")
    pdd = pdd.filter((pl.col('Date')>=primeiro)&(pl.col('Date')<=ultimo)).sort(by="Date")
    the_range = list(pl.date_range(primeiro, ultimo, eager=True))
    lista_segm = ['ACC_Total', 'ACC_Consumo', 'ACC_Moche', 'ACC_Uzo', 'ACC_Altice_Empresas', 'ACC_ACE', 'ACE_Total', 'ACE_Empresariais', 'ACE_Altice_Empresas', 'Altice_Empresas_Total']
    for i, dia in enumerate(the_range):
        lista=pdd.filter(pl.col('Date')==dia)
        val=pcd.filter(pl.col('Date')==dia)['New_Users'].to_list()[0]
        ws.cell(row=row_actual, column=coluna+i, value=val).style=normal
        for j, seg in enumerate(lista_segm):
            if (j+row_actual) < 30:
                ws.cell(row=row_actual+j, column=coluna+i, value=0 if lista[seg].shape[0]==0 else lista[seg].to_list()[0]).style=normal
            else:
                ws.cell(row=row_actual+j, column=coluna+i, value=0 if lista[seg].shape[0]==0 else lista[seg].to_list()[0]).style=normalunder
    return len(the_range)

def month_total(dados: dict, ws: worksheet, coluna: int, row_actual: int)->None:
    ultimo = last_day_of_last_month(datetime.today()).date()
    primeiro = datetime(ultimo.year, ultimo.month, 1).date()
    ante_ultimo = edate(datetime(datetime.today().year, datetime.today().month,1) -timedelta(days=1))
    ante_primeiro = datetime(ante_ultimo.year, ante_ultimo.month, 1).date()
    pcd = dados['new_users_day'].sort(by='Date')
    pcl = pcd.filter((pl.col('Date')>=primeiro)&(pl.col('Date')<=ultimo))
    pce = pcd.filter((pl.col('Date')>=ante_primeiro)&(pl.col('Date')<=ante_ultimo))
    pdd = dados['users_ac_day'].sort(by="Date")
    pdl = pdd.filter((pl.col('Date')>=primeiro)&(pl.col('Date')<=ultimo))
    pde = pdd.filter((pl.col('Date')>=ante_primeiro)&(pl.col('Date')<=ante_ultimo))
    lista_segm = ['ACC_Total', 'ACC_Consumo', 'ACC_Moche', 'ACC_Uzo', 'ACC_Altice_Empresas', 'ACC_ACE', 'ACE_Total', 'ACE_Empresariais', 'ACE_Altice_Empresas', 'Altice_Empresas_Total']
    ws.cell(row=row_actual-2, column=coluna, value='').style=normalgray
    ws.cell(row=row_actual-2, column=coluna+1, value='').style=normalgray
    ws.cell(row=row_actual-2, column=coluna+2, value='').style=normalgray
    ws.cell(row=row_actual-1, column=coluna, value=pcl['New_Users'].sum()).style=normalgray
    ws.cell(row=row_actual-1, column=coluna+1, value=pce['New_Users'].sum()).style=normalgray
    ws.cell(row=row_actual-1, column=coluna+2, value=pcl['New_Users'].sum()/pce['New_Users'].sum() - 1).style=normalgrayperc
    for j, seg in enumerate(lista_segm):
        if (j+row_actual) < 30:
            ws.cell(row=row_actual+j, column=coluna, value=pdl[seg].sum()).style=normalgray
            ws.cell(row=row_actual+j, column=coluna+1, value=pde[seg].sum()).style=normalgray
            ws.cell(row=row_actual+j, column=coluna+2, value=pdl[seg].sum()/pde[seg].sum()-1).style=normalgrayperc
        else:
            ws.cell(row=row_actual+j, column=coluna, value=pdl[seg].sum()).style=normalgrayunder
            ws.cell(row=row_actual+j, column=coluna+1, value=pde[seg].sum()).style=normalgrayunder
            ws.cell(row=row_actual+j, column=coluna+2, value=pdl[seg].sum()/pde[seg].sum()-1).style=normalgrayunderperc


def coluna_ytd_semanal_logins(
    ident_de_dados: str,
    switch_final: str,
    fifth_year: int,
    fifth_week: int,
    coluna: int,
    dados: dict,
    ws: worksheet,
    lista_unica: list,
    lista: list,
)->None:
    #wb.add_named_style(tlinhagrayperc) 
    #wb.add_named_style(normalgrayperc)
    #wb.add_named_style(normalgrayunderperc)
    lista_ytd_sem_unique = [
        sum(
            list(
                dados[ident_de_dados].filter(
                    (
                        dados[ident_de_dados]['Year']==fifth_year
                    )&(
                        dados[ident_de_dados]['Week']<fifth_week
                    )
                )[item]
            )
        ) for item in lista_unica
    ]

    lista_ytd_total=[sum(lista_ytd_sem_unique)]+lista_ytd_sem_unique[:2]+[
        sum(list(dados[ident_de_dados].filter((
            dados[ident_de_dados]['Year']==fifth_year
        )&(
    dados[ident_de_dados]['Week']<fifth_week))['Qtd_Logins_Uzo'])
        )
    ]+lista_ytd_sem_unique[-2:]
    num_ytd = [get_data_year_week(dados, ident_de_dados, fifth_year, fifth_week, item) for item in lista]
    num_ytd_total = [sum(num_ytd[:-1])]+num_ytd
    coluna_ytd = [
        last_week_ytd / sum_of_ytd - 1 if sum_of_ytd>0 else 0 for last_week_ytd, sum_of_ytd in zip(
            num_ytd_total,
            [tot/(fifth_week-1) if fifth_week>1 else 0 for tot in lista_ytd_total]
        )
    ]

    for line, val in enumerate(coluna_ytd):
        if line == 0:
            ws.cell(row=6+line, column=coluna+2, value=val).style=tlinhagrayperc
        else:
            ws.cell(row=6+line, column=coluna+2, value=val).style=normalgrayperc
    if switch_final!='Qtd_Logins_Unique':
        return
    numer_ult_racio = get_data_year_week(dados, ident_de_dados, fifth_year, fifth_week, 'Qtd_Logins_Unique' )
    denom_ulti_racio = sum(list(dados['logins_week'].filter((dados['logins_week']['Year']==fifth_year )&(dados['logins_week']['Week']<fifth_week))['Qtd_Logins_Unique']))
    ultimo_racio = numer_ult_racio / (denom_ulti_racio/(fifth_week-1)) - 1 if fifth_week>1 else 0
    ws.cell(row=12, column=coluna+2, value=ultimo_racio).style=normalgrayperc
    ws.cell(row=13, column=coluna+2, value='').style=normalgrayunderperc

def weeks_in_year(lista_pares: list[list[int, int]], linha: int, coluna: int, ws: worksheet, dados: dict[str, pl.DataFrame], indice: str, ind_total: str )->None:
    lista_segm = [
        'ACC_Total',
        'ACC_Consumo',
        'ACC_Moche',
        'ACC_Uzo',
        'ACC_Altice_Empresas',
        'ACC_ACE',
        'ACE_Total',
        'ACE_Empresariais',
        'ACE_Altice_Empresas',
        'Altice_Empresas_Total']
    tab_week = dados[indice]
    tab_ind_tot = dados[ind_total]
    for stp, (ano, week) in enumerate(lista_pares):
        col = coluna + stp
        reex = tab_ind_tot.filter((pl.col('Year')==ano)&(pl.col('Week')==week))
        temp = reex["New_Users"]
        for ind, titulo in enumerate(lista_segm):
            row = linha + ind
            reex = tab_week.filter((pl.col('Year')==ano)&(pl.col('Week')==week))
            temp = reex[titulo]
            valor = list(temp)[0]
            ws.cell(row=row, column=col, value=valor).style=normal
            
        

#def coluna_ytd_mensal_new_acc(
    
#):
def week_summary_1(
    lista_pares: list[list[int, int]], 
    linha: int, 
    coluna: int, 
    ws: worksheet, 
    dados: dict[str, pl.DataFrame], 
    indice: str, 
    ind_total: str ):
    lista_segm = [
        'ACC_Total',
        'ACC_Consumo',
        'ACC_Moche',
        'ACC_Uzo',
        'ACC_Altice_Empresas',
        'ACC_ACE',
        'ACE_Total',
        'ACE_Empresariais',
        'ACE_Altice_Empresas',
        'Altice_Empresas_Total']
    tab_week = dados[indice]
    tab_ind_tot = dados[ind_total]
    for stp, (ano, mes) in enumerate(lista_pares):
        col = coluna + stp
        reex = tab_ind_tot.filter((pl.col('Year')==ano)&(pl.col('Week')==mes))
        temp = reex["New_Users"]
        for ind, titulo in enumerate(lista_segm):
            row = linha + ind
            reex = tab_week.filter((pl.col('Year')==ano)&(pl.col('Week')==mes))
            temp = reex[titulo]
            valor = list(temp)[0]
            ws.cell(row=row, column=col, value=valor).style=normal
    
def months_in_year(
    ano: int,
    primeiro_mes: int,
    primeiro_ano: int,
    linha: int,
    coluna: int,
    dados: dict[str, pl.DataFrame],
    ws: worksheet
)->None:
    lista_segm = [
        'ACC_Total',
        'ACC_Consumo',
        'ACC_Moche',
        'ACC_Uzo',
        'ACC_Altice_Empresas',
        'ACC_ACE',
        'ACE_Total',
        'ACE_Empresariais',
        'ACE_Altice_Empresas',
        'Altice_Empresas_Total']
    dados_df = dados["users_ac_month"]
    total = 0
    ano: int = primeiro_ano
    mes: int = primeiro_mes
    for _ in range(12):
        for lis in range(len(lista_segm)):
            item: int = lis + linha
            temp1 = dados_df.filter((pl.col("Year")==ano) & (pl.col("Month")==mes))
            valor = list(temp1[lista_segm[lis]])[0]   
            total += valor 
            ws.cell(row=item, column=mes, value = valor).style=normal
        if mes + 1 > 12:
            mes = 1
            ano += 1
        else:
            mes += 1 
        ws.cell(row=20, column=mes, value = total).style=normal
    for _ in range(12):
        dados["new_users_month"]
        


