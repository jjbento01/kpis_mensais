import polars as pl
from formatos import *
from utilities import *
from openpyxl import worksheet
from datetime import datetime

def cabecalho_tv_logins(ws: worksheet, coluna: int, row_actual: int):
    tags = [ 
        'Logins TV', 
        '      Visits', 
        '      Intentional visits (> 10 seg)', 
        '      Unique visitors',
        '      Average visits per visitor']
    estilos = [total, normal, normal, normal, normal]
    for i, tag, sty in zip(range(0, 14), tags[:-1], estilos):
        ws.cell(row=row_actual+i, column=coluna,  value=tag).style=sty
    ws.cell(row=row_actual+i+1, column=coluna, value=tags[-1]).style=normalunder

def valores_diarios_tv_logins(dados: dict, ws: worksheet, coluna: int, row_actual: int):
    dados['logins_day_tv'] = dados['logins_day_tv'].sort(by="Date")
    for i, row in enumerate(dados['logins_day_tv'].iter_rows()):
        visits=0
        unique=0
        for j, celula in enumerate(row):
            if j > 0:
                ws.cell(row=row_actual+j-1, column=coluna+i, value=celula).style=normal
            if j==1:
                visits=celula
            if j==3:
                unique=celula
        ws.cell(row=row_actual+j, column=coluna+i, value=visits/unique if unique!=0 else 0).style=normalunder

def total_mes_analise(dados: dict, ws: worksheet, coluna: int, row_actual: int, primeiro: datetime, ultimo: datetime)->list:
    ultimo_mes_df = dados['logins_day_tv'].filter((pl.col('Date')>=primeiro.date())&(pl.col('Date')<=ultimo.date())).sort(by="Date")
    ws.cell(row=row_actual-1, column=coluna, value="").style=normalgray
    ws.cell(row=row_actual, column=coluna, value=ultimo_mes_df["Qtd_Visits"].sum()).style=normalgray
    ws.cell(row=row_actual+1, column=coluna, value=ultimo_mes_df["Qtd_Visits_Intentional"].sum()).style=normalgray
    ws.cell(row=row_actual+2, column=coluna, value=ultimo_mes_df["Qtd_Visits_Unique"].sum()).style=normalgray
    ws.cell(row=row_actual+3, column=coluna, value="").style=normalgrayunder
    return [ultimo_mes_df["Qtd_Visits"].sum(), ultimo_mes_df["Qtd_Visits_Intentional"].sum(), ultimo_mes_df["Qtd_Visits_Unique"].sum()]

def total_mes_analise_var(ws: worksheet, coluna: int, row_actual: int, lista_fll: list, lista_fla: list)->None:
    ws.cell(row=row_actual-1, column=coluna, value="").style=normalgray
    ws.cell(row=row_actual, column=coluna, value=lista_fll[0]/lista_fla[0]-1).style=normalgrayperc
    ws.cell(row=row_actual+1, column=coluna, value=lista_fll[1]/lista_fla[1]-1).style=normalgrayperc
    ws.cell(row=row_actual+2, column=coluna, value=lista_fll[2]/lista_fla[2]-1).style=normalgrayperc
    ws.cell(row=row_actual+3, column=coluna, value="").style=normalgrayunder
    
def calculate_week_columns(
    dados: dict, 
    tabela: str,
    unique_row: str,
    ws: worksheet, 
    coluna: int, 
    date_range: pl.Series,
    first_year: int,
    first_week: int,
    second_year: int,
    second_week: int,
    third_year: int,
    third_week: int,
    fourth_year: int,
    fourth_week: int,
    fifth_year: int,
    fifth_week: int,
    lista_unica: list, 
    row_totalinha: int,
    row_normalunder: int)->tuple:
    valores_totais: list[int] = []
    per_user_list: list[float] = []
    for coluna, (year, week) in zip(
            range(7, 12),
            [(first_year, first_week), 
            (second_year, second_week), 
            (third_year, third_week),
            (fourth_year, fourth_week),
            (fifth_year, fifth_week)]):
            temp = 0
            for i, item in enumerate(lista_unica):
                #temp += get_data_year_week(dados, tabela, year, week, item)
                ws.cell(row=row_totalinha+i, column=len(date_range)+coluna, value=get_data_year_week(dados, tabela, year, week, item)).style=nmgrds
            ws.cell(row=row_totalinha+len(lista_unica), column=len(date_range)+coluna, value=get_data_year_week(dados, tabela, year, week, unique_row)).style=nmgrds
            valores_totais.append(temp)
            #per_user= temp/get_data_year_week(dados, tabela, year, week, unique_row)
            #per_user_list.append(per_user)
            #ws.cell(row=row_totalinha, column=len(date_range)+coluna, value=temp).style=totalinha
            nom = get_data_year_week(dados, tabela, year, week, lista_unica[0])
            den = get_data_year_week(dados, tabela, year, week, unique_row)
            ws.cell(row=row_normalunder+1, column=len(date_range)+coluna, value=(nom/den) if den!=0 else 0).style=normalunder
    return valores_totais, per_user_list

def get_w48vs1w(dados: dict, tabela: str, lista: list, ws: worksheet, fifth_year: int, fifth_week: int, fourth_year: int, fourth_week: int, row: int, coluna: int):
    for i, item in enumerate(lista):
        numerador = get_data_year_week(dados, tabela, fifth_year, fifth_week, item)
        denominador = get_data_year_week(dados, tabela, fourth_year, fourth_week, item)
        ws.cell(row=row+i, column=coluna, value=(numerador/denominador) if denominador!=0 else 0-1).style=normalgrayperc
    nom1 = get_data_year_week(dados, tabela, fifth_year, fifth_week, lista[0])
    den1 = get_data_year_week(dados, tabela, fifth_year, fifth_week, lista[-1])
    nomt1 = (nom1/den1) if den1!=0 else 0
    nom2 = get_data_year_week(dados, tabela, fourth_year, fourth_week, lista[0])
    den2 = get_data_year_week(dados, tabela, fourth_year, fourth_week, lista[-1])
    nomt2 = (nom2/den2) if den2!=0 else 0
    ws.cell(row=row-1, column=coluna, value='').style=normalgray
    ws.cell(row=row+len(lista), column=coluna, value=((nomt1/nomt2) if nomt2!=0 else 0) -1).style=normalgrayunderperc

def get_w48vs4w(dados: dict, tabela: str, lista: list, ws: worksheet, 
                fifth_year: int, fifth_week: int, 
                fourth_year: int, fourth_week: int, 
                third_year: int, third_week: int, 
                second_year: int, second_week: int, 
                first_year: int, first_week: int, 
                row: int, coluna: int):
    racio=0
    for i, item in enumerate(lista):
        temp = 0
        for [year, week] in [[fourth_year, fourth_week], [third_year, third_week], [second_year, second_week], [first_year, first_week]]:
            temp += get_data_year_week(dados, tabela, year, week, item)
            if i==0:
                numerador = get_data_year_week(dados, tabela, year, week, lista[0])
                denominador = get_data_year_week(dados, tabela, year, week, lista[-1])
                tmp = (numerador/denominador) if denominador!=0 else 0
                racio += tmp
        numerador = get_data_year_week(dados, tabela, fifth_year, fifth_week, item)
        denominador = temp/4
        ws.cell(row=row+i, column=coluna, value=((numerador/denominador) if denominador!=0 else 0) -1).style=normalgrayperc
    ws.cell(row=row-1, column=coluna, value='').style=normalgray
    nom = get_data_year_week(dados, tabela, fifth_year, fifth_week, lista[0])
    den = get_data_year_week(dados, tabela, fifth_year, fifth_week, lista[-1])
    numerador = (nom/den) if den!=0 else 0
    ws.cell(row=row+len(lista), column=coluna, value=((numerador/(racio/4)) if racio!=0 else 0)-1).style=normalgrayunderperc
        
def tv_ytd_semanal_web_logins(ident_de_dados: str, fifth_year: int, fifth_week: int, coluna: int, dados: dict, ws: worksheet, row: int, lista: list)-> None:
    lista_ytd = [
        sum(
            list(
                dados[ident_de_dados].filter(
                    (
                        dados[ident_de_dados]['Year']==fifth_year
                    )&(
                        dados[ident_de_dados]['Week']<fifth_week
                    )
                )[item]
            )
        ) for item in lista
    ]
    #retlst = []
    ws.cell(row=row-1, column=coluna, value='').style=normalgrayperc
    for i, item in enumerate(lista):
        num = get_data_year_week(dados, ident_de_dados, fifth_year, fifth_week, item)
        den = (fifth_week-1)
        valor = num/den if den!=0 else 0
        ws.cell(row=row+i, column=coluna, value=valor-1).style=normalgrayperc
    ws.cell(row=row+len(lista), column=coluna, value='').style=normalgrayunder
    
def colunas_mensais_tv_logins(dados: dict, tabela: str, lista: list, ws: worksheet, linha: int, coluna:int)->list:
    pda=dados[tabela].with_columns(("0"+pl.col('Month').cast(pl.String)).str.slice(-2).alias('daystr'))
    pdb = pda.with_columns(pl.col('Year').cast(pl.String).alias('ordyear'))
    pdc = pdb.with_columns(pl.concat_str([pl.col('ordyear'), pl.col('daystr')], separator="").alias('ord'))
    pdd = pdc.sort(by="ord")
    retorno = []
    for mes, mes_row in enumerate(pdd.iter_rows()):
        #primeira_coluna=sum([mes_row[col] for col in range(2,5)])
        ws.cell(row=linha-1, column=coluna+mes, value='').style=totalinha
        nom = 0
        den = 0
        for col, val in enumerate(mes_row):
            if col > 1 and col <= len(lista)+1:
                ws.cell(row=linha+col-2, column=coluna+mes, value=val).style=nmgrds
            if col== 2:
                nom = val
            if col== 4:
                den = val
        ws.cell(row=linha+len(lista), column=coluna+mes, value=(nom/den) if den!=0 else 0).style=normalunder
        retorno.append((nom/den) if den!=0 else 0)
    return retorno

def coluna_mensal_4_meses_tv(dados: dict, ws: worksheet, coluna: int, lista: list, lista_lg_us: list) -> None:
    ultimo_dia_ultimo_mes = datetime(datetime.now().year, datetime.now().month, 1) - timedelta(days=1)
    ano_a_um_mes = ultimo_dia_ultimo_mes.year
    mes_a_um_mes = ultimo_dia_ultimo_mes.month
    ano_corrente = ano_a_um_mes
    mes_corrente = mes_a_um_mes
    lista_acumula = [0]*len(lista)
    for _ in range(4):
        data_anterior = datetime(ano_corrente, mes_corrente, 1) - timedelta(days=1)
        ano_corrente = data_anterior.year
        mes_corrente = data_anterior.month
        for i, item in enumerate(lista):
            lista_acumula[i]+=dados['logins_month_tv'].filter((pl.col('Year')==ano_corrente)&(pl.col('Month')==mes_corrente))[item].to_list()[0]
    lista_acumula_total = [item/4.0 for item in lista_acumula]
    lista_atual=[dados['logins_month_tv'].filter((pl.col('Year')==ano_a_um_mes)&(pl.col('Month')==mes_a_um_mes))[item].to_list()[0] for item in lista]
    total_menos_um = sum(lista_acumula_total[:-1])
    total_atual = sum(lista_atual[:-1])
    ws.cell(row=15, column=coluna, value=total_atual/total_menos_um-1).style=tlinhagrayperc
    for line, num, den in zip(range(5), lista_atual, lista_acumula_total):
        ws.cell(row=16+line, column=coluna, value=num/den-1).style=normalgrayperc
    num0 = lista_atual[-1]
    den0 = lista_acumula_total[-1]
    num1 = lista_lg_us[-1]
    den1 = (sum(lista_lg_us[-5:-1])/4)
    ws.cell(row=20, column=coluna, value=((num0/den0) if den0!=0 else 0) - 1).style=normalgrayperc  
    ws.cell(row=21, column=coluna, value=((num1/den1) if den1!=0 else 0) - 1).style=normalgrayunderperc

def tv_make_last_month(dados: dict, ws: worksheet, coluna: int, lista: list, lista_rets: list)->None:
    ultimo_dia_ultimo_mes = datetime(datetime.now().year, datetime.now().month, 1) - timedelta(days=1)
    ano_a_um_mes = ultimo_dia_ultimo_mes.year
    mes_a_um_mes = ultimo_dia_ultimo_mes.month
    ultimo_dia_penult_mes = datetime(ultimo_dia_ultimo_mes.year, ultimo_dia_ultimo_mes.month, 1) - timedelta(days=1)
    ano_a_dois_meses = ultimo_dia_penult_mes.year
    mes_a_dois_meses = ultimo_dia_penult_mes.month
    ws.cell(row=15-1, column=coluna, value='').style=normalgrayperc
    for i, item in enumerate(lista):
        if item==lista[-1]:
            ws.cell(row=15+i, column=coluna, value=(lista_rets[-1]/lista_rets[-2])-1).style=normalgrayunderperc
        else:
            mes_anterior=get_data_year_month(dados, 'logins_month_tv', ano_a_um_mes, mes_a_um_mes, item)
            mes_penultim=get_data_year_month(dados, 'logins_month_tv', ano_a_dois_meses, mes_a_dois_meses, item)
            ws.cell(row=15+i, column=coluna, value=mes_anterior/mes_penultim-1).style=normalgrayperc

def tv_make_last_four_month(dados: dict, ws: worksheet, coluna: int, lista: list, lista_rets: list)->None:
    ws.cell(row=15-1, column=coluna, value='').style=normalgrayperc
    for i, item in enumerate(lista):
        ultimo_dia_ultimo_mes = datetime(datetime.now().year, datetime.now().month, 1) - timedelta(days=1)
        ano_a_um_mes = ultimo_dia_ultimo_mes.year
        mes_a_um_mes = ultimo_dia_ultimo_mes.month
        acm = 0
        denominador = get_data_year_month(dados, 'logins_month_tv', ano_a_um_mes, mes_a_um_mes, item)
        for _ in range(4):
            itera_mes = datetime(ano_a_um_mes, mes_a_um_mes, 1) - timedelta(days=1)
            ano_a_um_mes = itera_mes.year
            mes_a_um_mes = itera_mes.month
            acm += get_data_year_month(dados, 'logins_month_tv', ano_a_um_mes, mes_a_um_mes, item)
        media = acm/4
        ws.cell(row=15+i, column=coluna, value=denominador/media-1).style=normalgrayperc
    media = sum(lista_rets[-5:-1])/4
    ws.cell(row=15+len(lista), column=coluna, value=lista_rets[-1]/media-1).style=normalgrayunderperc
    
def coluna_ytd_mensal_tv_logins(
    ident_de_dados: str,
    fifth_year: int, 
    fifth_month: int, 
    coluna: int, 
    dados: dict, 
    ws: worksheet, 
    lista: list,
)->None:
    #wb.add_named_style(tlinhagrayperc)
    #wb.add_named_style(normalgrayperc)
    #wb.add_named_style(normalgrayunderperc)
    ws.cell(row=15-1, column=coluna, value='').style=normalgrayperc
    for i, item in enumerate(lista):
        soma = sum(
            list(
                dados[ident_de_dados].filter(
                    dados[ident_de_dados]['Year']*100+dados[ident_de_dados]['Month'] < (fifth_year*100+fifth_month)
                )[item]
            )
        )
        denom = get_data_year_month(dados,'logins_month_tv',fifth_year,fifth_month,item)
        ws.cell(row=15+i, column=coluna, value=denom/(soma/((fifth_month-1) if fifth_month!=1 else 12))-1).style=normalgrayperc
    ws.cell(row=15+len(lista), column=coluna, value='').style=normalgrayunderperc

def tv_coluna_vazia(ws: worksheet, linha: int, coluna: int, lista: list)->None:
    ws.cell(row=linha-1, column=coluna, value='').style=normalgray
    for i, item in enumerate(lista):
        if i == len(lista):
            ws.cell(row=linha+i, column=coluna, value='').style=normalgrayunderperc
        ws.cell(row=linha+i, column=coluna, value='').style=normalgray
        
