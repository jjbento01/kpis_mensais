
from openpyxl import Workbook
from datetime import datetime, timedelta
from openpyxl import worksheet
import polars as pl
from formatos import *
from utilities import *


def soma_de_logins_web(lista_unica: list, dicio: dict, data: datetime) -> int:
    """
    Calculate the sum of logins for a specific date.

    Args:
        lista_unica (list[str]): A list of login type column names to sum.
        dicio (dict[str, pl.DataFrame]): A dictionary containing DataFrames with login data.
        data (datetime): The date for which to calculate the sum of logins.

    Returns:
        int: The sum of logins for the specified date across the given columns.
    """
    soma: int = 0
    for i in lista_unica:
        soma += dicio["logins_day"].filter(pl.col("Date") == pl.lit(data))[i].item()
    return soma


def average_logins_per_user(lista_unica: list, dicio: dict, data: datetime) -> float:
    """
    Calculate the average logins per user for a specific date.

    Args:
        lista_unica (list[str]): A list of login type column names to calculate the average.
        dicio (dict[str, pl.DataFrame]): A dictionary containing DataFrames with login data.
        data (datetime): The date for which to calculate the average.

    Returns:
        float: The average logins per user for the specified date across the given columns.
    """
    return round(
        soma_de_logins_web(
            lista_unica,
            dicio,
            data
        )/dicio["logins_day"].filter(
            pl.col("Date") == pl.lit(data)
        )["Qtd_Logins_Unique"].item(),1)

def coluna_ytd_semanal_web_logins(
    ident_de_dados: str,
    switch_final: str,
    fifth_year: int,
    fifth_week: int,
    coluna: int,
    dados: dict,
    ws: worksheet,
    lista_unica: list,
    lista: list,
)->None:
    #wb.add_named_style(tlinhagrayperc)
    #wb.add_named_style(normalgrayperc)
    #wb.add_named_style(normalgrayunderperc)
    lista_dias = []
    for item in dados[ident_de_dados].iter_rows(named=True):
        lista_dias.append(get_first_day_of_week(item["Year"], item["Week"]))
    dados[ident_de_dados] = dados[ident_de_dados].with_columns(pl.Series("Data", lista_dias).cast(pl.Date))
    dados[ident_de_dados] = dados[ident_de_dados].sort(by="Data", descending=True)
    ano=dados[ident_de_dados][0]["Data"][0].year
    week=dados[ident_de_dados][0]["Data"][0].isocalendar().week
    #min_year = datetime(ano,1,1)-timedelta(days=datetime(ano,1,1).isocalendar().weekday-1)
    lista_ytd_sem_unique = dados[ident_de_dados].filter((pl.col("Year")==ano)&(pl.col("Week")<week)).select(lista_unica)
    lista0 = dados[ident_de_dados].filter((pl.col("Year")==ano)&(pl.col("Week")==week)).select(lista_unica)
    #lista_ytd_sem_unique_soma = lista_ytd_sem_unique.with_columns(sum(lista_unica).alias("total"))
    lista_ytd_sem_unique_soma = lista_ytd_sem_unique.with_columns(pl.fold(0, lambda acc, s: acc + s, pl.all()).alias("sum_horizontal"))
    lista0_soma = lista0.with_columns(pl.fold(0, lambda acc, s: acc + s, pl.all()).alias("sum_horizontal"))
    valor_total = lista0_soma["sum_horizontal"][0]/(lista_ytd_sem_unique_soma["sum_horizontal"].sum()/lista_ytd_sem_unique_soma.shape[0])-1
    lista_unique = dados[ident_de_dados].filter((pl.col("Year")==ano)&(pl.col("Week")<week))
    lista_unique0 = dados[ident_de_dados].filter((pl.col("Year")==ano)&(pl.col("Week")==week))
    ws.cell(row=6, column=coluna+2, value=valor_total).style=tlinhagrayperc
    for i, item in enumerate(lista_unica):
        denom = lista_ytd_sem_unique_soma[item].sum()
        valor_total = lista0_soma[item][0]/(denom/lista_ytd_sem_unique_soma.shape[0])-1 if denom!=0 else ''
        ws.cell(row=7+i, column=coluna+2, value=valor_total).style=normalgrayperc
    valor_total = lista_unique0[lista[-1]][0]/(lista_unique[lista[-1]].sum()/lista_unique.shape[0])-1
    ws.cell(row=12, column=coluna+2, value=valor_total).style=normalgrayperc
    ws.cell(row=13, column=coluna+2, value='').style=normalgrayunderperc

def colunas_mensais_web_logins(dados: dict, ws: worksheet, coluna:int)->list:
    pda=dados['logins_month'].with_columns(("0"+pl.col('Month').cast(pl.String)).str.slice(-2).alias('daystr'))
    pdb = pda.with_columns(pl.col('Year').cast(pl.String).alias('ordyear'))
    pdc = pdb.with_columns(pl.concat_str([pl.col('ordyear'), pl.col('daystr')], separator="").alias('ord'))
    pdd = pdc.sort(by="ord")
    retorno = []
    for mes, row in enumerate(pdd.iter_rows()):
        primeira_coluna=sum([row[col] for col in range(2,7)])
        dianome = datetime(row[0], row[1],1).strftime("%b %y")
        ws.cell(row=2, column=coluna+mes, value='').style=totalinha
        ws.cell(row=3, column=coluna+mes, value=dianome).style=monthyearsimple
        ws.cell(row=6, column=coluna+mes, value=primeira_coluna).style=totalinha
        for col, val in enumerate(row):
            if col > 1 and col < 8:
                ws.cell(row=5+col, column=coluna+mes, value=val).style=nmgrds
        ws.cell(row=13, column=coluna+mes, value=primeira_coluna/row[7]).style=normalunder
        retorno.append(primeira_coluna/row[7])
    return retorno

def coluna_mensal_um_mes(dados: dict, ws: worksheet, coluna: int, lista: list) -> None:
    ultimo_dia_ultimo_mes = datetime(datetime.now().year, datetime.now().month, 1) - timedelta(days=1)
    ano_a_um_mes = ultimo_dia_ultimo_mes.year
    mes_a_um_mes = ultimo_dia_ultimo_mes.month
    ultimo_dia_penultimo_mes = datetime(ano_a_um_mes, mes_a_um_mes, 1) - timedelta(days=1)
    ano_penultimo_mes = ultimo_dia_penultimo_mes.year
    mes_penultimo_mes = ultimo_dia_penultimo_mes.month
    lista_nums_este = [dados['logins_month'].filter((pl.col('Year')==ano_a_um_mes)&(pl.col('Month')==mes_a_um_mes))[item].to_list()[0] for item in lista]
    lista_nums_pass = [dados['logins_month'].filter((pl.col('Year')==ano_penultimo_mes)&(pl.col('Month')==mes_penultimo_mes))[item].to_list()[0] for item in lista]
    resultado_inter = [(este/passo-1 if passo!=0 else 0) for este, passo in zip(lista_nums_este, lista_nums_pass)]
    soma_ultimo = sum(lista_nums_este[:-1])
    soma_penult = sum(lista_nums_pass[:-1])
    ws.cell(row=2, column=coluna, value=ultimo_dia_ultimo_mes.strftime("%b %y")+ " F vs.").style=tlinhagrayperc
    ws.cell(row=3, column=coluna, value="1m").style=normalgrayunderperc
    ws.cell(row=6, column=coluna, value=soma_ultimo/soma_penult-1).style=tlinhagrayperc
    for line, val in enumerate(resultado_inter):
        ws.cell(row=7+line, column=coluna, value=val).style=normalgrayperc
    nume = soma_ultimo/lista_nums_este[-1]
    deno = soma_penult/lista_nums_pass[-1]
    ws.cell(row=13, column=coluna, value=nume/deno-1).style=normalgrayunderperc

def coluna_mensal_4_meses(dados: dict, ws: worksheet, coluna: int, lista: list, lista_lg_us: list) -> None:
    ultimo_dia_ultimo_mes = datetime(datetime.now().year, datetime.now().month, 1) - timedelta(days=1)
    ano_a_um_mes = ultimo_dia_ultimo_mes.year
    mes_a_um_mes = ultimo_dia_ultimo_mes.month
    ano_corrente = ano_a_um_mes
    mes_corrente = mes_a_um_mes
    lista_acumula = [0]*len(lista)

    for _ in range(4):
        data_anterior = datetime(ano_corrente, mes_corrente, 1) - timedelta(days=1)
        ano_corrente = data_anterior.year
        mes_corrente = data_anterior.month
        for i, item in enumerate(lista):
            lista_acumula[i]+=dados['logins_month'].filter((pl.col('Year')==ano_corrente)&(pl.col('Month')==mes_corrente))[item].to_list()[0]

    lista_acumula_total = [item/4.0 for item in lista_acumula]
    lista_atual=[dados['logins_month'].filter((pl.col('Year')==ano_a_um_mes)&(pl.col('Month')==mes_a_um_mes))[item].to_list()[0] for item in lista]
    total_menos_um = sum(lista_acumula_total[:-1])
    total_atual = sum(lista_atual[:-1])
    ws.cell(row=2, column=coluna, value="").style=tlinhagrayperc
    ws.cell(row=3, column=coluna, value="Avg.4m").style=normalgrayunderperc
    ws.cell(row=6, column=coluna, value=total_atual/total_menos_um-1).style=tlinhagrayperc
    for line, num, den in zip(range(5), lista_atual, lista_acumula_total):
        ws.cell(row=7+line, column=coluna, value=num/den-1).style=normalgrayperc
    ws.cell(row=12, column=coluna, value=lista_atual[-1]/lista_acumula_total[-1]-1).style=normalgrayperc
    ws.cell(row=13, column=coluna, value=lista_lg_us[-1]/(sum(lista_lg_us[-5:-1])/4)-1).style=normalgrayunderperc

def coluna_vazia(ws: worksheet, coluna: int):
    ws.cell(row=2, column=coluna, value="").style=tlinhagraypercenter
    ws.cell(row=3, column=coluna, value="Avg.YTD").style=normalgrayunderpercenter
    ws.cell(row=6, column=coluna, value="").style=tlinhagrayperc
    for line in range(5):
        ws.cell(row=7+line, column=coluna, value="").style=normalgrayperc
    ws.cell(row=12, column=coluna, value="").style=normalgrayperc
    ws.cell(row=13, column=coluna, value="").style=normalgrayunderperc

def coluna_vazia_fim(ws: worksheet, coluna: int, ultimo:datetime):
    ws.cell(row=2, column=coluna, value="AsIs").style=tlinhagraypercenter
    ws.cell(row=3, column=coluna, value=ultimo.strftime("%y/%b")).style=normalgrayunderpercenter
    ws.cell(row=6, column=coluna, value="").style=tlinhagrayperc
    for line in range(5):
        ws.cell(row=7+line, column=coluna, value="").style=normalgrayperc
    ws.cell(row=12, column=coluna, value="").style=normalgrayperc
    ws.cell(row=13, column=coluna, value="").style=normalgrayunderperc

def coluna_ytd_mensal_logins(dados: dict, ws: worksheet, coluna: int, lista: list):
    ultimo_dia_ultimo_mes = datetime(datetime.now().year, datetime.now().month, 1) - timedelta(days=1)
    pda=dados["logins_month"].filter((pl.col('Year')==ultimo_dia_ultimo_mes.year)&(pl.col('Month')<ultimo_dia_ultimo_mes.month))
    lista_totais = [list(list(pda.select(pl.nth(i)).sum())[0])[0]/(12 if (ultimo_dia_ultimo_mes.month-1)==0 else (ultimo_dia_ultimo_mes.month - 1)) for i in range(2,8)]
    linhas_atual = [item[0] for item in list(dados['logins_month'].filter((pl.col('Year')==ultimo_dia_ultimo_mes.year)&(pl.col('Month')==ultimo_dia_ultimo_mes.month)))][2:]
    valores=[0 if media==0 else atual/media-1 for atual, media in zip(linhas_atual, lista_totais)]
    ws.cell(row=2, column=coluna, value="").style=tlinhagrayperc
    ws.cell(row=3, column=coluna, value=ultimo_dia_ultimo_mes.strftime("%d/%b")).style=normalgrayunderpercenter
    val = ((sum(linhas_atual[:-1])/sum(lista_totais[:-1])) if sum(lista_totais[:-1])!=0 else 0) - 1
    ws.cell(row=6, column=coluna, value=val).style=tlinhagrayperc
    for line, val in enumerate(valores):
        ws.cell(row=7+line, column=coluna, value=val).style=normalgrayperc
    ws.cell(row=13, column=coluna, value="").style=normalgrayunderperc

def bloco_principal_logins(
    ws: worksheet,
    first_last_month: datetime,
    last_last_month: datetime,
    primeiro_ultimo_mes: datetime,
    ultimo: datetime,
    lista_unica: list,
    date_range: list,
    dados: dict,
    tags: list,
    estilos: list):
    for i, tag, sty in zip(range(4, 14), tags, estilos):
        ws.cell(row=i, column=2,  value=tag).style=sty
    logins_total=0
    logins_total_1=0
    logins_my_meo=0
    logins_my_meo_1=0
    logins_moche=0
    logins_moche_1=0
    logins_uzo=0
    logins_uzo_1=0
    logins_mypte=0
    logins_mypte_1=0
    logins_acpte=0
    logins_acpte_1=0
    logins_unique=0
    logins_unique_1=0
    for i,data in enumerate(date_range):
        #ic(data)
        ws.cell(row=2, column=3+i, value=day_of_week[data.weekday()]).style=dayweek
        ws.cell(row=3, column=3+i, value=((data - datetime(1899, 12, 30).date()).days)).style=normalshort
        ws.cell(row=13, column=3+i, value=average_logins_per_user(lista_unica, dados, data)).style=normalunder
        ws.column_dimensions[ws.cell(row=5, column=3+i).column_letter].width=13
        temp = dados["logins_day"].filter( pl.col("Date") == pl.lit(data) )['Qtd_Logins_MyMeo'][0]
        soma = temp
        logins_my_meo += temp if data>=primeiro_ultimo_mes.date() and data<=ultimo.date() else 0
        logins_my_meo_1 += temp if data>=first_last_month.date() and data<=last_last_month.date() else 0
        ws.cell(row=7, column=3+i, value=temp).style=nmgrds
        temp = dados["logins_day"].filter( pl.col("Date") == pl.lit(data) )['Qtd_Logins_Moche'][0]
        soma += temp
        logins_moche += temp if data>=primeiro_ultimo_mes.date() and data<=ultimo.date() else 0
        logins_moche_1 += temp if data>=first_last_month.date() and data<=last_last_month.date() else 0
        ws.cell(row=8, column=3+i, value=temp).style=nmgrds
        temp = dados["logins_day"].filter( pl.col("Date") == pl.lit(data) )['Qtd_Logins_Uzo'][0]
        soma += temp
        logins_uzo += temp if data>=primeiro_ultimo_mes.date() and data<=ultimo.date() else 0
        logins_uzo_1 += temp if data>=first_last_month.date() and data<=last_last_month.date() else 0
        ws.cell(row=9, column=3+i, value=temp).style=nmgrds
        temp = dados["logins_day"].filter(pl.col("Date") == pl.lit(data))['Qtd_Logins_MyPTE'][0]
        soma += temp
        logins_mypte += temp if data>=primeiro_ultimo_mes.date() and data<=ultimo.date() else 0
        logins_mypte_1 += temp if data>=first_last_month.date() and data<=last_last_month.date() else 0
        ws.cell(row=10, column=3+i, value=temp).style=nmgrds
        temp = dados["logins_day"].filter(pl.col("Date") == pl.lit(data))['Qtd_Logins_ACE'][0]
        soma += temp
        logins_acpte += temp if data>=primeiro_ultimo_mes.date() and data<=ultimo.date() else 0
        logins_acpte_1 += temp if data>=first_last_month.date() and data<=last_last_month.date() else 0
        ws.cell(row=11, column=3+i, value=temp).style=nmgrds
        temp = dados["logins_day"].filter(pl.col("Date") == pl.lit(data))['Qtd_Logins_Unique'][0]
        logins_unique += temp if data>=primeiro_ultimo_mes.date() and data<=ultimo.date() else 0
        logins_unique_1 += temp if data>=first_last_month.date() and data<=last_last_month.date() else 0
        ws.cell(row=6, column=3+i, value=soma).style=totalinha
        ws.cell(row=12, column=3+i, value=temp).style=nmgrds
        logins_total+=soma if data>=primeiro_ultimo_mes.date() and data<=ultimo.date() else 0
        logins_total_1+=soma if data>=first_last_month.date() and data<=last_last_month.date() else 0

    for i, label1, label2 in zip(range(3, 6), ((ultimo - datetime(1899, 12, 30)), "", ""), ("Total", "M-1", "MoM")):
        ws.cell(row=2, column=len(date_range)+i, value=label1).style=monthyear
        ws.cell(row=3, column=len(date_range)+i, value=label2).style=normalgrayunder
        ws.cell(row=13, column=len(date_range)+i, value="").style=normalgrayunder

    ws.column_dimensions[ws.cell(row=2, column=len(date_range)+6).column_letter].width=3
    for ipos, sty, vals in zip(range(6, 13),
                        [[totalinhagray, totalinhagray, tlinhagrayperc],
                        [normalgray, normalgray, normalgrayperc],
                        [normalgray, normalgray, normalgrayperc],
                        [normalgray, normalgray, normalgrayperc],
                        [normalgray, normalgray, normalgrayperc],
                        [normalgray, normalgray, normalgrayperc],
                        [normalgray, normalgray, normalgrayperc]],
                        [[logins_total, logins_total_1, logins_total/logins_total_1-1],
                        [logins_my_meo, logins_my_meo_1, logins_my_meo/logins_my_meo_1-1],
                        [logins_moche, logins_moche_1, logins_moche/logins_moche_1-1],
                        [logins_uzo, logins_uzo_1, logins_uzo/logins_uzo_1-1],
                        [logins_mypte, logins_mypte_1, logins_mypte/logins_mypte_1-1 if logins_mypte_1!=0 else 0],
                        [logins_acpte, logins_acpte_1, logins_acpte/logins_acpte_1-1 if logins_acpte_1!=0 else 0],
                        [logins_unique, logins_unique_1, logins_unique/logins_unique_1-1 if logins_unique_1!=0 else 0]]):
        for jpos, esc, valor in zip(range(3, 6), sty, vals):
            ws.cell(row=ipos, column=len(date_range)+jpos, value=valor).style=esc
