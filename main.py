import pandas as pd
import concurrent.futures
import threading
from yaml import load, FullLoader
import codecs
import os
from sqlalchemy import engine, create_engine
from datetime import datetime, timedelta
#import pandas as pd
import polars as pl
from openpyxl import Workbook
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import NamedStyle, Side
from icecream import ic


from lista_queries import lista, lista_unica, lista_tags_users, lista_segm, lista_iteracoes, lista_queries_a_fazer, tags_logins
from utilities import get_first_day_of_week, calc_date_values, range_needed_of_date, bloco_principal_logins, get_data_year_week, coluna_ytd_semanal_web_logins, colunas_mensais_web_logins, coluna_mensal_um_mes, coluna_mensal_4_meses, coluna_ytd_mensal_logins, coluna_vazia, cabecalho_tv_logins, valores_diarios_tv_logins, total_mes_analise, edate, total_mes_analise_var, calculate_week_columns, get_w48vs1w, get_w48vs4w, tv_ytd_semanal_web_logins, colunas_mensais_tv_logins, tv_make_last_month, tv_make_last_four_month, tv_coluna_vazia, novos_users_ac_cabecalhos, valores_diarios_new_users, valores_diarios_users, month_total, acrescentar_colunas_semanais, week_summary_1, months_in_year, month_summary_1, final_values
from formatos import cabecalho, total, normal, normalshort, totalinha, totalinhadir, nmgrds, nmgrdsund, normalgrayperc, normalgrayunder, tlinhagrayperc, normalalgray, normalgrayunderpercenter, nmgrdseuro, nmgrdsundeuro, nmgrdseurogray, nmgrdsundeurogray, normalgrayunderperc
from logins import get_data_year_week, getdays_for_week, get_w48vs1w, get_w48vs4w, tv_ytd_semanal_web_logins, colunas_mensais_tv_logins, tv_make_last_month, tv_make_last_four_month, coluna_ytd_mensal_tv_logins, tv_coluna_vazia
from logins_tv import *
from novos_users_ac import *
from pandas.core.indexes.datetimes import date_range



with codecs.open(str(os.getenv("CONFS"))+"conf_dba\\configuracao.yaml", "r", "utf-8") as fich:
    config = load(fich, Loader=FullLoader)
    string_conf = config['con_str_0']

bd: str ='BD_ECARE'
engine: engine = create_engine(f'mssql+pyodbc:///?odbc_connect={ string_conf };DATABASE='+bd)
dados: dict[str, pl.DataFrame]={}

def download_site(query: list[str,str])->None:
    global engine, dados
    ic(query)
    dados[query[0]] = pl.read_database(query[1], engine)

def download_all_sites(lista_queries: list[list[str,str]])->None:
    with concurrent.futures.ThreadPoolExecutor(max_workers=7) as executor:
        executor.map(download_site, lista_queries)

def do_main()->tuple:
    global lista_queries_a_fazer, lista, lista_unica,dados
    download_all_sites(lista_queries_a_fazer)
    lista_dias = []
    for item in dados["users_ac_week"].iter_rows(named=True):
        lista_dias.append(get_first_day_of_week(item["Year"], item["Week"]))
    dados["users_ac_week"] = dados["users_ac_week"].with_columns(pl.Series("Data", lista_dias).cast(pl.Date))
    #dados["users_ac_week"] = dados["users_ac_week"].with_columns(pl.struct(["Year", "Week"]).map_batches(lambda x: get_first_day_of_week(x["Year"], x["Week"])).alias("Date"))
    dados["users_ac_week"] = dados["users_ac_week"].sort(by="Data", descending=True)
    cinco_semanas = dados["users_ac_week"].head(5)
    first_day = cinco_semanas["Data"][0]
    first_year = cinco_semanas['Year'][0]
    first_week = cinco_semanas['Week'][0]
    second_day = cinco_semanas["Data"][1]
    second_year = cinco_semanas['Year'][1]
    second_week = cinco_semanas['Week'][1]
    third_day = cinco_semanas["Data"][2]
    third_year = cinco_semanas['Year'][2]
    third_week = cinco_semanas['Week'][2]
    fourth_day = cinco_semanas["Data"][3]
    fourth_year = cinco_semanas['Year'][3]
    fourth_week = cinco_semanas['Week'][3]
    fifth_day = cinco_semanas["Data"][4]
    fifth_year = cinco_semanas['Year'][4]
    fifth_week = cinco_semanas['Week'][4]
    date_range = range_needed_of_date(datetime.today())
    [ultimo, _, primeiro_ultimo_mes, first_last_month, last_last_month] = calc_date_values()
    #[[first_day, first_year, first_week],
    # [second_day, second_year, second_week],
    # [third_day, third_year, third_week],
    # [fourth_day, fourth_year, fourth_week],
    # [fifth_day, fifth_year, fifth_week]] = getdays_for_week(dados)
    ic(date_range[0].strftime("%Y-%m-%d"))
    ic(primeiro_ultimo_mes)
    ic(date_range[-1].strftime("%Y-%m-%d"))
    ic(first_last_month.strftime("%Y-%m-%d"))
    ic(last_last_month.strftime("%Y-%m-%d"))
    ic("=====================================")
    wb = Workbook()
    #carregar_format(wb)
    ws = wb.active
    ws.title="Report"
    ws.column_dimensions['A'].width=5
    ws.column_dimensions['B'].width=30
    
    estilos = [cabecalho, cabecalho, total, normal, normal, normal, normal, normal, normal, normal]
    bloco_principal_logins(ws, first_last_month, last_last_month, primeiro_ultimo_mes, ultimo, lista_unica, date_range, dados, tags_logins, estilos)
    # linha com o cabeçalho das semanas
    for coluna, iweek, iday in zip(range(7,13), (fifth_week, fourth_week, third_week, second_week, first_week), (fifth_day, fourth_day, third_day, second_day, first_day)): 
        ws.cell(row=2, column=len(date_range)+coluna, value="W"+str(iweek)).style=totalinhadir
        ws.cell(
            row=3,
            column=len(date_range)+coluna,
            value=(datetime(iday.year, iday.month, iday.day, 0,0,0) - datetime(1899, 12, 30)).days-1
        ).style=normalshort
    medias: dict[str, float] = {item: 0 for item in lista}
    for linha, item in zip([7, 8, 9, 10, 11, 12], lista):
        for ipos, (year, week) , sty in zip(range(7, 13), [(fifth_year, fifth_week), (fourth_year, fourth_week), (third_year, third_week), (second_year, second_week),(first_year, first_week)], [nmgrds, nmgrds, nmgrds, nmgrds, nmgrds]):
            temp = get_data_year_week(dados, 'logins_week', year, week, item)
            ws.cell(row=linha, column=len(date_range)+ipos, value=temp).style=sty
            #import ipdb; ipdb.set_trace()
            medias[item]+=temp if not(year==first_year and week==first_week) else 0

    per_user_list: list[float] = []

    valores_totais: list[float] = []
    lista_semanas = [
        (first_year, first_week),
        (second_year, second_week),
        (third_year, third_week),
        (fourth_year, fourth_week),
        (fifth_year, fifth_week)
    ]
    lista_semanas.reverse()
    # constroi a coluna das semanas
    # dos logins web, sem ser os tv logins os logins tv são mais adiante
    for coluna, (year, week) in zip(
        range(7, 12),
        lista_semanas):
        temp = 0
        # item é cada um dos membros da lista de categorias, por exemplo: "Qtd_Logins_MyMeo"
        for item in lista_unica:
            temp += get_data_year_week(dados, 'logins_week', year, week, item)
        valores_totais.append(temp)
        per_user= temp/get_data_year_week(dados, 'logins_week', year, week, 'Qtd_Logins_Unique')
        per_user_list.append(per_user)
        ws.cell(row=6, column=len(date_range)+coluna, value=temp).style=totalinha
        ws.cell(row=13, column=len(date_range)+coluna, value=per_user).style=normalunder

    coluna: int+p = len(date_range)+12

    total_first: int = 0
    total_second: int = 0

    for linha, item, color in zip(
        range(len(lista_unica)),
        lista_unica,
        [
            normalgrayperc,
            normalgrayperc,
            normalgrayperc,
            normalgrayperc,
            normalgrayperc,
            normalgrayperc
        ]):
        temp_first: int = get_data_year_week(dados, 'logins_week', first_year, first_week, item)
        temp_second: int = get_data_year_week(dados, 'logins_week', second_year, second_week, item)
        rto: float = 0 if temp_second == 0 else temp_first / temp_second - 1
        ws.cell(row=7+linha, column=coluna, value=rto).style=color
        total_first += temp_first
        total_second += temp_second

    resultado: float = get_data_year_week(dados,
        'logins_week',
        first_year,
        first_week,
        'Qtd_Logins_Unique'
    )/get_data_year_week(dados,
        'logins_week',
        second_year,
        second_week,
        'Qtd_Logins_Unique') - 1

    ic(resultado)

    ws.cell(row=7+len(lista_unica), column=coluna, value=resultado).style=normalgrayperc

    ws.cell(row=6, column=coluna, value=total_first/total_second-1).style=tlinhagrayperc
    ws.cell(row=13, column=coluna, value=per_user_list[-1]/per_user_list[-2]-1).style=normalgrayunderperc

    for i, primeiralinha, segundalinha in zip(range(3), ("W"+str(fifth_week)+" vs.", "", ""), ("1w", "Avg.4w", "Avg.YTD")):
        ws.cell(row=2, column=coluna+i, value=primeiralinha).style=tlinhagrayperc
        ws.cell(row=3, column=coluna+i, value=segundalinha).style=normalgrayunder

    valores_inter = [dados['logins_week'].filter((dados['logins_week']['Year']==first_year)&(dados['logins_week']['Week']==first_week))[item]/(medias[item]/4)-1 for item in lista]
    total_val = valores_totais[-1]/(sum([medias[key] for key in medias.keys()][:-1])/4)-1
    ws.cell(row=6, column=coluna+1, value=total_val).style=tlinhagrayperc
    for i, valor in enumerate(valores_inter):
        ws.cell(row=7+i, column=coluna+1, value=list(valor)[0]).style=normalgrayperc
    total_fim = per_user_list[-1]/(sum(per_user_list[:-1])/4)-1
    ws.cell(row=13, column=coluna+1, value=total_fim).style=normalgrayunderperc
    coluna_ytd_semanal_web_logins('logins_week', "Qtd_Logins_Unique", fifth_year, fifth_week, coluna, dados, ws, lista_unica, lista)
    letra_no_fim = ws.cell(row=5, column=(len(date_range)-5)).column_letter
    ws.column_dimensions.group('C', letra_no_fim, hidden=True)
    ws.column_dimensions[ws.cell(row=2, column=coluna+3).column_letter].width=3
    coluna += 4
    coluna_inicial = coluna
    lista_logins_users = colunas_mensais_web_logins(dados, ws, coluna)
    coluna += 12
    letra_inicial = ws.cell(row=5, column=coluna_inicial).column_letter
    letra_final = ws.cell(row=5, column=coluna_inicial+7).column_letter
    ws.column_dimensions.group(letra_inicial, letra_final, hidden=True)
    coluna_mensal_um_mes(dados, ws, coluna, lista)
    coluna+=1
    coluna_mensal_4_meses(dados, ws, coluna, lista, lista_logins_users)
    coluna+=1
    coluna_ytd_mensal_logins(dados, ws, coluna, lista)
    coluna+=1
    ws.column_dimensions[ws.cell(row=2, column=coluna).column_letter].width=3
    coluna+=1
    coluna_vazia(ws,coluna)
    row_actual = 14
    coluna = 2
    cabecalho_tv_logins(ws, coluna, row_actual)
    row_actual = 15
    coluna += 1
    valores_diarios_tv_logins(dados, ws, coluna, row_actual)
    coluna += dados['logins_day_tv'].shape[0]
    primeiro_dia_mes = (datetime(datetime.now().year, datetime.now().month, 1) - timedelta(days=1)).replace(day=1)
    ultimo_dia_mes = datetime(datetime.now().year, datetime.now().month, 1) - timedelta(days=1)
    lista_fla = total_mes_analise(dados, ws, coluna, row_actual, primeiro_dia_mes, ultimo_dia_mes)
    coluna += 1
    primeiro_dia_anterior = (primeiro_dia_mes - timedelta(days=1)).replace(day=1)
    ultimo_dia_anterior = edate(ultimo_dia_mes)
    lista_fll = total_mes_analise(dados, ws, coluna, row_actual, primeiro_dia_anterior, ultimo_dia_anterior)
    coluna += 1
    total_mes_analise_var(ws, coluna, row_actual, lista_fla, lista_fll)
    coluna +=3
    calculate_week_columns(dados, 'logins_week_tv', 'Qtd_Visits_Unique', ws, coluna, date_range, fifth_year, fifth_week, fourth_year, fourth_week, third_year, third_week, second_year, second_week, first_year, first_week, ['Qtd_Visits', 'Qtd_Visits_Intentional'], 15, 17)
    coluna += 4
    get_w48vs1w(dados, "logins_week_tv", ['Qtd_Visits', 'Qtd_Visits_Intentional', 'Qtd_Visits_Unique'], ws, first_year, first_week, second_year, second_week, 15, coluna)
    coluna += 1
    get_w48vs4w(dados, 'logins_week_tv', ['Qtd_Visits', 'Qtd_Visits_Intentional', 'Qtd_Visits_Unique'], ws, first_year, first_week, second_year, second_week, third_year, third_week, fourth_year, fourth_week, fifth_year, fifth_week, 15, coluna)
    coluna += 1
    tv_ytd_semanal_web_logins('logins_week_tv', first_year, first_week, coluna, dados, ws, 15, ['Qtd_Visits', 'Qtd_Visits_Intentional', 'Qtd_Visits_Unique'])
    coluna += 2
    ret: list[float] = colunas_mensais_tv_logins(dados, 'logins_month_tv', ['Qtd_Visits', 'Qtd_Visits_Intentional', 'Qtd_Visits_Unique'], ws, 15, coluna)
    coluna += 12
    tv_make_last_month(dados, ws, coluna, ['Qtd_Visits', 'Qtd_Visits_Intentional', 'Qtd_Visits_Unique', ''], ret)
    coluna += 1
    tv_make_last_four_month(dados, ws, coluna, ['Qtd_Visits', 'Qtd_Visits_Intentional', 'Qtd_Visits_Unique'], ret)
    coluna += 1
    coluna_ytd_mensal_tv_logins('logins_month_tv', first_year, ultimo.month, coluna, dados, ws, ['Qtd_Visits', 'Qtd_Visits_Intentional', 'Qtd_Visits_Unique'])
    coluna += 2
    tv_coluna_vazia(ws, 15, coluna, ['Qtd_Visits', 'Qtd_Visits_Intentional', 'Qtd_Visits_Unique', ''])
    contador: int = 0
    # -------------------------------------------------------------------------------------------------------
    # começa linhas dos novos users ac
    # -------------------------------------------------------------------------------------------------------
    coluna = 2
    novos_users_ac_cabecalhos(ws, coluna, 20, lista_tags_users)
    coluna += 1
    valores_diarios_new_users("new_users_day", dados, ws, coluna, 20, nmgrds)
    coluna += valores_diarios_users(dados, ws, coluna, 21, lista_segm, 'new_users_day', 'users_ac_day', 30, nmgrds, nmgrds, nmgrdsund)
    month_total(dados, ws, coluna, 21, lista_segm, 'new_users_day', 'New_Users', 'users_ac_day', 30,
            normalalgray if contador <2 else nmgrdseurogray,
            normalgrayunder if contador <2 else nmgrdsundeurogray,
            normalgrayperc,
            normalgrayunderpercenter)
    for i in range(3): ws.cell(row=19, column=coluna+i, value='').style=normalgrayperc
    coluna += 4
    acrescentar_colunas_semanais(21, coluna, dados, ws, 'new_users_week', 'users_ac_week', 30, lista_segm, nmgrds if contador <2 else nmgrdseuro, nmgrdsund if contador<2 else nmgrdsundeuro)
    coluna += 5
    #month_col = coluna
    # coluna += 1
    week_summary_1([[fifth_year, fifth_week], [fourth_year, fourth_week], [third_year, third_week], [second_year, second_week], [first_year, first_week]], 20, coluna, ws, dados, "users_ac_week", "New_Users", "new_users_week", lista_segm, 29)
    for i in range(3): ws.cell(row=19, column=coluna+i, value='').style=normalgrayperc
    # month_col += 9
    coluna += 3
    ic(ultimo)
    primeiro_ano = (ultimo - pd.DateOffset(years=1) ) + timedelta(days=1)
    ic(primeiro_ano)
    months_in_year(fifth_year, primeiro_ano.month, first_year, 21, coluna, dados, ws, lista_segm, 30, "users_ac_month", "new_users_month", nmgrds if contador <2 else nmgrdseuro, nmgrdsund if contador<2 else nmgrdsundeuro)
    coluna += 13
    month_summary_1([[first_year, first_week], [second_year, second_week], [third_year, third_week], [fourth_year, fourth_week], [fifth_year, fifth_week]], 20, coluna, ws, dados, "users_ac_month", "new_users_month", "New_Users", lista_segm, 29)
    for i in range(3): ws.cell(row=19, column=coluna+i, value='').style=normalgrayperc
    coluna += 4
    linha = 20
    final_values(ultimo, linha, coluna, ws, dados, "users_ac_asis", "new_users_month", lista_segm)
    ws.cell(row=19, column=coluna, value='').style=normalgrayperc
    # acaba as linhas de novos users ac
    
    for composto in lista_iteracoes:
        [coluna, linha, fim, tags_iteracao, listagem] = composto
        fim += linha
        novos_users_ac_cabecalhos(ws, coluna, linha, tags_iteracao)
        coluna += 1
        coluna += valores_diarios_users(dados, ws, coluna, linha, listagem, None, 'cpag_day', fim, nmgrds if contador <2 else nmgrdseuro, nmgrds if contador <2 else nmgrdseuro, nmgrdsund if contador<2 else nmgrdsundeuro)
        month_total(dados, ws, coluna, linha, listagem, None, None, 'cpag_day', fim, normalalgray if contador <2 else nmgrdseurogray, normalgrayunder if contador <2 else nmgrdsundeurogray, normalgrayperc, normalgrayunderpercenter)
        coluna += 4
        acrescentar_colunas_semanais(linha, coluna, dados, ws, None, "cpag_week", fim, listagem, nmgrds if contador <2 else nmgrdseuro, nmgrdsund if contador<2 else nmgrdsundeuro)
        coluna += 5
        week_summary_1([[fifth_year, fifth_week], [fourth_year, fourth_week], [third_year, third_week], [second_year, second_week], [first_year, first_week]], linha-1, coluna, ws, dados, "cpag_week", None, None, listagem, fim-1)
        coluna += 3
        months_in_year(fifth_year, primeiro_ano.month, first_year, linha, coluna, dados, ws, listagem, fim, "cpag_month", None, nmgrds if contador <2 else nmgrdseuro, nmgrdsund if contador<2 else nmgrdsundeuro)
        coluna += 13
        month_summary_1([[fifth_year, fifth_week], [fourth_year, fourth_week], [third_year, third_week], [second_year, second_week], [first_year, first_week]], linha-1, coluna, ws, dados, "cpag_month", None, None, listagem, fim-1)
        coluna += 4
        for i in range(linha-1, fim+1): ws.cell(row=i, column=coluna, value='').style=(normalgrayperc if i < fim else normalgrayunderperc)
        contador += 1
        # fim das linhas dos cpag
        # Fim de Carregamentos por AC €

    wb.save("sample.xlsx")
    engine.dispose()
    return (dados, valores_inter, medias)
    
    
if __name__ == '__main__':
    dados, valores_inter, medias = do_main()
