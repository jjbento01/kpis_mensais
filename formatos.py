from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import NamedStyle, Side
from openpyxl import Workbook

bd = Side(style='thin', color="000000")

cabecalho = NamedStyle(name="cabecalho")
cabecalho.font = Font(name='Calibri',bold=True, size=9, color="FFFFFF")
cabecalho.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

cabelinha = NamedStyle(name="cabelinha")
cabelinha.font = Font(name='Calibri',bold=True, size=9, color="FFFFFF")
cabelinha.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
cabelinha.border = Border(top=Side(style='thin', color="FFFFFF"))

totalinha = NamedStyle(name="totalinha")
totalinha.font = Font(name='Calibri',bold=False, size=8, color="000000")
totalinha.border = Border(top=Side(style='thin', color="000000"))
totalinha.number_format = '#,##'

totalinhadir = NamedStyle(name="totalinhadir")
totalinhadir.font = Font(name='Calibri',bold=False, size=8, color="000000")
totalinhadir.border = Border(top=Side(style='thin', color="000000"))
totalinhadir.number_format = '#,##'
totalinhadir.alignment=Alignment(horizontal='right', vertical='center')

totalinhagray = NamedStyle(name="totalinhagray")
totalinhagray.font = Font(name='Calibri',bold=False, size=8, color="000000")
totalinhagray.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
totalinhagray.border = Border(top=Side(style='thin', color="000000"))
totalinhagray.number_format = '#,##'

tlinhagrayperc = NamedStyle(name="tlinhagrayperc")	
tlinhagrayperc.font = Font(name='Calibri',bold=False, size=8, color="000000")
tlinhagrayperc.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
tlinhagrayperc.border = Border(top=Side(style='thin', color="000000"))
tlinhagrayperc.number_format = '0.0%'

tlinhagraypercenter = NamedStyle(name="tlinhagraypercenter")	
tlinhagraypercenter.font = Font(name='Calibri',bold=False, size=8, color="000000")
tlinhagraypercenter.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
tlinhagraypercenter.border = Border(top=Side(style='thin', color="000000"))
totalinhadir.alignment=Alignment(horizontal='center', vertical='center')
tlinhagraypercenter.number_format = '0.0%'

highlight = NamedStyle(name="highlight")
highlight.font = Font(name='Calibri',bold=True, size=9)
highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
#highlight.number_format = 'yyyy-mm-dd'

total = NamedStyle(name="total")
total.font = Font(name='Calibri',bold=True, size=9)
total.border = Border(top=Side(style='thin', color="000000"))

dayweek = NamedStyle(name="dayweek")
dayweek.font = Font(name='Calibri',bold=False, size=8)
bd = Side(style='thin', color="000000")
dayweek.border = Border(top=bd)
dayweek.alignment=Alignment(horizontal='right', vertical='center')

monthyearsimple = NamedStyle(name="monthyearsimple")
monthyearsimple.font = Font(name='Calibri',bold=False, size=8)
bd = Side(style='thin', color="000000")
monthyearsimple.border = Border(bottom=bd)
monthyearsimple.alignment=Alignment(horizontal='right', vertical='center')

normaldata = NamedStyle(name="normaldata")
normaldata.font = Font(name='Calibri',bold=False, size=8)
normaldata.number_format = 'yyyy-mm-dd'

normalshort = NamedStyle(name="normalshort")
normalshort.font = Font(name='Calibri',bold=False, size=8)
normalshort.border = Border(bottom=Side(style='thin', color="000000"))
normalshort.number_format = '[$-409]dd Mmm'

monthyear = NamedStyle(name="monthyear")
monthyear.font = Font(name='Calibri',bold=False, size=8)
monthyear.border = Border(top=Side(style='thin', color="000000"))
monthyear.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
monthyear.alignment=Alignment(horizontal='center', vertical='center')
monthyear.number_format = '[$-409]Mmm yy'

normal = NamedStyle(name="normal")
normal.font = Font(name='Calibri',bold=False, size=8)
normal.number_format = '#,##0.0'

normalunder = NamedStyle(name="normalunder")
normalunder.font = Font(name='Calibri',bold=False, size=8)
normalunder.number_format = '#,##0.0'
normalunder.border = Border(bottom=Side(style='thin', color="000000"))

nmgrds = NamedStyle(name="nmgrds")
nmgrds.font = Font(name='Calibri',bold=False, size=8)
nmgrds.number_format = '#,##'   

normalgray = NamedStyle(name="normalgray")
normalgray.font = Font(name='Calibri',bold=False, size=8)
normalgray.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
normalgray.number_format = '#,##'   

normalgrayperc = NamedStyle(name="normalgrayperc")
normalgrayperc.font = Font(name='Calibri',bold=False, size=8)
normalgrayperc.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
normalgrayperc.number_format = '0.0%'   

normalalgray = NamedStyle(name="normalalgray")
normalalgray.font = Font(name='Calibri',bold=False, size=8)
normalalgray.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
normalalgray.alignment=Alignment(horizontal='right', vertical='center')
normalalgray.number_format = '#,##'   

normalgrayunder = NamedStyle(name="normalgrayunder")
normalgrayunder.font = Font(name='Calibri',bold=False, size=8)
normalgrayunder.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
normalgrayunder.alignment=Alignment(horizontal='right', vertical='center')
normalgrayunder.border = Border(bottom=Side(style='thin', color="000000"))
normalgrayunder.number_format = '#,##'   

normalgrayunderperc = NamedStyle(name="normalgrayunderperc")
normalgrayunderperc.font = Font(name='Calibri',bold=False, size=8)
normalgrayunderperc.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
normalgrayunderperc.alignment=Alignment(horizontal='right', vertical='center')
normalgrayunderperc.border = Border(bottom=Side(style='thin', color="000000"))
normalgrayunderperc.number_format = '0.0%'   

normalgrayunderpercenter = NamedStyle(name="normalgrayunderpercenter")
normalgrayunderpercenter.font = Font(name='Calibri',bold=False, size=8)
normalgrayunderpercenter.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
normalgrayunderpercenter.alignment=Alignment(horizontal='center', vertical='center')
normalgrayunderpercenter.border = Border(bottom=Side(style='thin', color="000000"))
normalgrayunderpercenter.number_format = '0.0%'   

def carregar_format(wb: Workbook)->None:
    wb.add_named_style(cabecalho)
    wb.add_named_style(cabelinha)
    wb.add_named_style(highlight)
    wb.add_named_style(normal)
    wb.add_named_style(cabecalho)
    wb.add_named_style(cabelinha)
    wb.add_named_style(nmgrds)
    wb.add_named_style(dayweek)
    wb.add_named_style(total)
    wb.add_named_style(totalinha)
    wb.add_named_style(normalshort)
    wb.add_named_style(monthyear)
    wb.add_named_style(totalinhagray)
    wb.add_named_style(normalalgray)
    wb.add_named_style(normalgrayunder)
    wb.add_named_style(tlinhagrayperc)
    wb.add_named_style(tlinhagraypercenter)
    wb.add_named_style(normalgrayperc)
    wb.add_named_style(normalunder)
    wb.add_named_style(totalinhadir)
    wb.add_named_style(monthyearsimple)
    wb.add_named_style(normalgrayunderpercenter)