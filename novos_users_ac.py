import polars as pl
from formatos import *
from utilities import *
from openpyxl import worksheet
from datetime import datetime, timedelta


# with codecs.open(str(os.getenv("CONFS"))+"conf_dba\\configuracao.yaml", "r", "utf-8") as fich:
#     config = load(fich, Loader=FullLoader)
#     string_conf = config['con_str_0']

# bd: str ='BD_ECARE'
# engine: engine = create_engine(f'mssql+pyodbc:///?odbc_connect={ string_conf };DATABASE='+bd)
# dados: dict[str, pl.DataFrame]={}

def novos_users_ac_cabecalhos(ws: worksheet, coluna: int, row_actual:int, tags: list[str])->None:
    estilos = [total, normal, normal, normal, normal, normal, normal, normal, normal, normal, normal, normal]
    for i, tag, sty in zip(range(0, 13), tags[:-1], estilos):
        ws.cell(row=row_actual+i-1, column=coluna,  value=tag).style=sty
    ws.cell(row=row_actual+i, column=coluna, value=tags[-1]).style=normalunder

def valores_diarios_new_users(tabela: str, dados: dict, ws: worksheet, coluna: int, row_actual: int)->None:
    primeiro = first_day_a_year_ago(datetime.today().date())
    ultimo = last_day_of_last_month(datetime.today()).date()
    pdd = dados[tabela].filter((pl.col('Date')>=primeiro)&(pl.col('Date')<=ultimo)).sort(by="Date")
    for i, row in enumerate(pdd.iter_rows()):
        ws.cell(row=row_actual, column=coluna+i, value=row[1]).style=normal

def valores_diarios_users(dados: dict, ws: worksheet, coluna: int, row_actual: int, lista_segm: list[str], topo: str, corpo: str, limite_linhas: int)->int:
    primeiro = first_day_a_year_ago(datetime.today().date())
    ultimo = last_day_of_last_month(datetime.today()).date()
    if topo is not None:
        pcd = dados[topo].sort(by='Date')
    pdd = dados[corpo].sort(by="Date")
    pdd = pdd.filter((pl.col('Date')>=primeiro)&(pl.col('Date')<=ultimo)).sort(by="Date")
    the_range = list(pl.date_range(primeiro, ultimo, eager=True))
    #lista_segm = ['ACC_Total', 'ACC_Consumo', 'ACC_Moche', 'ACC_Uzo', 'ACC_Altice_Empresas', 'ACC_ACE', 'ACE_Total', 'ACE_Empresariais', 'ACE_Altice_Empresas', 'Altice_Empresas_Total']
    for i, dia in enumerate(the_range):
        lista=pdd.filter(pl.col('Date')==dia)
        if topo is not None:
            val=pcd.filter(pl.col('Date')==dia)['New_Users'].to_list()[0]
            ws.cell(row=row_actual, column=coluna+i, value=val).style=normal
        for j, seg in enumerate(lista_segm):
            if (j+row_actual) < limite_linhas:
                ws.cell(row=row_actual+j, column=coluna+i, value=0 if lista[seg].shape[0]==0 else lista[seg].to_list()[0]).style=nmgrds
            else:
                ws.cell(row=row_actual+j, column=coluna+i, value=0 if lista[seg].shape[0]==0 else lista[seg].to_list()[0]).style=nmgrdsund
    return len(the_range)

def month_total(dados: dict, ws: worksheet, coluna: int, row_actual: int, lista_segm: list[str], head: str, fld_head: str, body: str, limite_maximo: int)->None:
    ultimo = last_day_of_last_month(datetime.today()).date()
    primeiro = datetime(ultimo.year, ultimo.month, 1).date()
    ante_ultimo = edate(datetime(datetime.today().year, datetime.today().month,1) -timedelta(days=1))
    ante_primeiro = datetime(ante_ultimo.year, ante_ultimo.month, 1).date()
    if head is not None:
        pcd = dados[head].sort(by='Date')
        pcl = pcd.filter((pl.col('Date')>=primeiro)&(pl.col('Date')<=ultimo))
        pce = pcd.filter((pl.col('Date')>=ante_primeiro)&(pl.col('Date')<=ante_ultimo))
    pdd = dados[body].sort(by="Date")
    pdl = pdd.filter((pl.col('Date')>=primeiro)&(pl.col('Date')<=ultimo))
    pde = pdd.filter((pl.col('Date')>=ante_primeiro)&(pl.col('Date')<=ante_ultimo))
    ws.cell(row=row_actual-1, column=coluna, value='').style=normalgray
    ws.cell(row=row_actual-1, column=coluna+1, value='').style=normalgray
    ws.cell(row=row_actual-1, column=coluna+2, value='').style=normalgray
    if head is not None:
        ws.cell(row=row_actual-1, column=coluna, value=pcl[fld_head].sum()).style=normalgray
        ws.cell(row=row_actual-1, column=coluna+1, value=pce[fld_head].sum()).style=normalgray
        ws.cell(row=row_actual-1, column=coluna+2, value=pcl[fld_head].sum()/pce[fld_head].sum() - 1).style=normalgrayperc
    for j, seg in enumerate(lista_segm):
        if (j+row_actual) < limite_maximo:
            ws.cell(row=row_actual+j, column=coluna, value=pdl[seg].sum()).style=normalgray
            ws.cell(row=row_actual+j, column=coluna+1, value=pde[seg].sum()).style=normalgray
            ws.cell(row=row_actual+j, column=coluna+2, value=((pdl[seg].sum()/pde[seg].sum() - 1) if pde[seg].sum()>0 else '')).style=normalgrayperc
        else:
            ws.cell(row=row_actual+j, column=coluna, value=pdl[seg].sum()).style=normalgrayunder
            ws.cell(row=row_actual+j, column=coluna+1, value=pde[seg].sum()).style=normalgrayunder
            ws.cell(row=row_actual+j, column=coluna+2, value=((pdl[seg].sum()/pde[seg].sum() - 1) if pde[seg].sum()>0 else '')).style=normalgrayunderperc

def acrescentar_colunas_semanais(inicio: int, coluna: int, dados: dict[str, pl.DataFrame], ws: worksheet, head: str, body: str, limite_maximo: int, lista_segm: list[str])->None:
    if head is not None:
        df_escolhida_new: pl.DataFrame=dados[head].sort(by=['Year', 'Week']).drop(['Year', 'Week']).tail(5)
    df_escolhida_acc: pl.DataFrame=dados[body].sort(by=['Year', 'Week']).drop(['Year', 'Week']).tail(5)
    for i, row in enumerate(lista_segm):
        for j, col in enumerate(df_escolhida_acc[row].to_list()):
            if (i+inicio) < limite_maximo:
                ws.cell(row=inicio+i, column=coluna+j, value=col).style=nmgrds
            else:
                ws.cell(row=inicio+i, column=coluna+j, value=col).style=nmgrdsund
    if head is not None:
        for i, row in enumerate(df_escolhida_new.iter_rows()):
            ws.cell(row=inicio-1, column=coluna+i, value=row[0]).style=nmgrds
        
def coluna_ytd_semanal_logins(
    ident_de_dados: str,
    switch_final: str,
    fifth_year: int,
    fifth_week: int,
    coluna: int,
    dados: dict,
    ws: worksheet,
    lista_unica: list,
    lista: list,
)->None:
    #wb.add_named_style(tlinhagrayperc) 
    #wb.add_named_style(normalgrayperc)
    #wb.add_named_style(normalgrayunderperc)
    lista_ytd_sem_unique = [
        sum(
            list(
                dados[ident_de_dados].filter(
                    (
                        dados[ident_de_dados]['Year']==fifth_year
                    )&(
                        dados[ident_de_dados]['Week']<fifth_week
                    )
                )[item]
            )
        ) for item in lista_unica
    ]

    lista_ytd_total=[sum(lista_ytd_sem_unique)]+lista_ytd_sem_unique[:2]+[
        sum(list(dados[ident_de_dados].filter((
            dados[ident_de_dados]['Year']==fifth_year
        )&(
            dados[ident_de_dados]['Week']<fifth_week))['Qtd_Logins_Uzo'])
        )
    ]+lista_ytd_sem_unique[-2:]
    num_ytd = [get_data_year_week(dados, ident_de_dados, fifth_year, fifth_week, item) for item in lista]
    num_ytd_total = [sum(num_ytd[:-1])]+num_ytd
    coluna_ytd = [
        last_week_ytd / sum_of_ytd - 1 if sum_of_ytd>0 else 0 for last_week_ytd, sum_of_ytd in zip(
            num_ytd_total,
            [tot/(fifth_week-1) if fifth_week>1 else 0 for tot in lista_ytd_total]
        )
    ]

    for line, val in enumerate(coluna_ytd):
        if line == 0:
            ws.cell(row=6+line, column=coluna+2, value=val).style=tlinhagrayperc
        else:
            ws.cell(row=6+line, column=coluna+2, value=val).style=normalgrayperc
    if switch_final!='Qtd_Logins_Unique':
        return
    numer_ult_racio = get_data_year_week(dados, ident_de_dados, fifth_year, fifth_week, 'Qtd_Logins_Unique' )
    denom_ulti_racio = sum(list(dados['logins_week'].filter((dados['logins_week']['Year']==fifth_year )&(dados['logins_week']['Week']<fifth_week))['Qtd_Logins_Unique']))
    ultimo_racio = numer_ult_racio / (denom_ulti_racio/(fifth_week-1)) - 1 if fifth_week>1 else 0
    ws.cell(row=12, column=coluna+2, value=ultimo_racio).style=normalgrayperc
    ws.cell(row=13, column=coluna+2, value='').style=normalgrayunderperc
        

def week_summary_1(
    lista_pares: list[list[int, int]], 
    linha: int, 
    coluna: int, 
    ws: worksheet, 
    dados: dict[str, pl.DataFrame], 
    indice: str, 
    ind_acc: str,
    ind_total: str,
    lista_segm: list[str],
    limite_maximo: int)->None:
    if ind_total is not None:
        primeira_coluna_denom = dados[ind_total].filter((pl.col('Year')==lista_pares[4][0])&(pl.col('Week')==lista_pares[4][1]))
        primeira_coluna_numer = dados[ind_total].filter((pl.col('Year')==lista_pares[3][0])&(pl.col('Week')==lista_pares[3][1]))   
        perc = primeira_coluna_denom[ind_acc]/primeira_coluna_numer[ind_acc] - 1.0
        ws.cell(row=linha, column=coluna, value=perc[0]).style=normalgrayperc
        soma_denom = sum([dados[ind_total].filter((pl.col('Year')==lista_pares[3][0])&(pl.col('Week')==lista_pares[3][1]))[ind_acc] for i in range(4)])
        perc = primeira_coluna_denom[ind_acc]/(soma_denom / 4.0) - 1.0
        ws.cell(row=linha, column=coluna+1, value=perc[0]).style=normalgrayperc
        total_denom = dados[ind_total].filter((pl.col('Year')==lista_pares[4][0])&(pl.col('Week')<lista_pares[4][1]))[ind_acc].sum()
        perc = primeira_coluna_denom[ind_acc]/total_denom
        ws.cell(row=linha, column=coluna+2, value=perc[0]).style=normalgrayperc
    for i, item in enumerate(lista_segm):
        primeira_coluna_denom = dados[indice].filter((pl.col('Year')==lista_pares[4][0])&(pl.col('Week')==lista_pares[4][1]))[item]
        primeira_coluna_numer = dados[indice].filter((pl.col('Year')==lista_pares[3][0])&(pl.col('Week')==lista_pares[3][1]))[item]
        perc = primeira_coluna_denom/primeira_coluna_numer - 1.0
        ws.cell(row=linha+i+1, column=coluna, value=perc[0]).style=(normalgrayperc if (i+linha) < limite_maximo else normalgrayunderperc)
        soma_denom=sum([dados[indice].filter((pl.col('Year')==lista_pares[i][0])&(pl.col('Week')==lista_pares[i][1]))[item] for i in range(4)])
        perc = (float(primeira_coluna_denom[0])/(float(soma_denom[0])/4.0) - 1.0) if soma_denom[0]>0 else ''
        ws.cell(row=linha+i+1, column=coluna+1, value=perc).style=(normalgrayperc if (i+linha) < limite_maximo else normalgrayunderperc)
        total_denom = dados[indice].filter((pl.col('Year')==lista_pares[4][0])&(pl.col('Week')<lista_pares[4][1]))[item].sum()
        perc = primeira_coluna_denom[0]/total_denom if total_denom>0 else ''
        ws.cell(row=linha+i+1, column=coluna+2, value=perc).style=(normalgrayperc if (i+linha) < limite_maximo else normalgrayunderperc)

def months_in_year(
    ano: int,
    primeiro_mes: int,
    primeiro_ano: int,
    linha: int,
    coluna: int,
    dados: dict[str, pl.DataFrame],
    ws: worksheet,
    lista_segm: list[str],
    listagem_maximo: int,
    ident: str,
    ident_local: str
)->None:
    # lista_segm = [
    #     'ACC_Total',
    #     'ACC_Consumo',
    #     'ACC_Moche',
    #     'ACC_Uzo',
    #     'ACC_Altice_Empresas',
    #     'ACC_ACE',
    #     'ACE_Total',
    #     'ACE_Empresariais',
    #     'ACE_Altice_Empresas',
    #     'Altice_Empresas_Total']
    dados_df = dados[ident]
    dados_df = dados_df.with_columns((pl.col("Year").cast(pl.Utf8) + "-" + pl.col("Month").cast(pl.Utf8) + "-01").str.strptime(pl.Date, "%Y-%m-%d").alias("Data"))
    dados_df = dados_df.sort(by=["Data"])
    ano: int = primeiro_ano
    mes: int = primeiro_mes
    for i, row in enumerate(dados_df.iter_rows()):
        lista_row = {k: v for k, v in zip(dados_df.columns, row)}
        for i, seg in enumerate(lista_segm): 
            ws.cell(row=i+linha, column=(mes+coluna), value = lista_row[seg]).style=(nmgrds if linha+i<listagem_maximo else nmgrdsund)
        if mes + 1 > 12:
            mes = 1
            ano += 1
        else:
            mes += 1 
        #ws.cell(row=20, column=(mes+coluna), value = total).style=normal
    if ident_local is not None:
        for i, row in enumerate(dados[ident_local].sort(by=["Year", "Month"])[-12:].iter_rows()):
            ws.cell(row=20, column=(coluna +  i+ 1), value = row[2]).style=(nmgrds if linha+i<listagem_maximo else nmgrdsund)
        
def month_summary_1(
    lista_pares: list[list[int, int]], 
    linha: int, 
    coluna: int, 
    ws: worksheet, 
    dados: dict[str, pl.DataFrame], 
    indice: str, 
    ind_total: str,
    ind_col: str,
    lista_segm: list[str],
    listagem_maximo: int)->None:
    """
    Escreve resumo de dados de New_Users por semana no m s corrente e nos 3 meses anteriores.
    Parameters
    ----------
    lista_pares : list[list[int, int]]
        Lista de pares de ano e m s para o resumo.
    linha : int
        Linha onde o resumo ser  escrito.
    coluna : int
        Coluna onde o resumo ser  escrito.
    ws : worksheet
        Worksheet onde o resumo ser  escrito.
    dados : dict[str, pl.DataFrame]
        Dicion rio com os dados a serem escritos.
    indice : str
        Nome do dicion rio com os dados a serem escritos.
    ind_total : str
        Nome do dicion rio com os dados totais a serem escritos.
    ind_col : str
        Nome da coluna de dados a serem escritos.
    lista_segm : list[str]
        Lista com os nomes dos dados a serem escritos.
    listagem_maximo : int
        N mero m ximo de linhas a serem escritas.
    """
    if ind_total is not None:
        dados[ind_total] = dados[ind_total].with_columns((pl.col("Year").cast(pl.Utf8) + "-" + pl.col("Month").cast(pl.Utf8) + "-01").str.strptime(pl.Date, "%Y-%m-%d").alias("Data"))
    dados[indice] = dados[indice].with_columns((pl.col("Year").cast(pl.Utf8) + "-" + pl.col("Month").cast(pl.Utf8) + "-01").str.strptime(pl.Date, "%Y-%m-%d").alias("Data"))
    month_ultimo = dados[indice].sort(by=["Data"])[-1]["Month"][0]
    year_ultimo = dados[indice].sort(by=["Data"])[-1]["Year"][0]
    month_penultimo = dados[indice].sort(by=["Data"])[-2]["Month"][0]
    year_penultimo = dados[indice].sort(by=["Data"])[-2]["Year"][0]
    month_antepenultimo = dados[indice].sort(by=["Data"])[-3]["Month"][0]
    year_antepenultimo = dados[indice].sort(by=["Data"])[-3]["Year"][0]
    month_anteantepenultimo = dados[indice].sort(by=["Data"])[-4]["Month"][0]
    year_anteantepenultimo = dados[indice].sort(by=["Data"])[-4]["Year"][0]
    month_primeiro = dados[indice].sort(by=["Data"])[-5]["Month"][0]
    year_primeiro = dados[indice].sort(by=["Data"])[-5]["Year"][0]
    lista_anos = [year_penultimo, year_antepenultimo, year_anteantepenultimo, year_primeiro]
    lista_meses = [month_penultimo, month_antepenultimo, month_anteantepenultimo, month_primeiro]
    ultima_data = dados[indice].sort(by=["Data"])[-1]["Data"][0]
    if ind_total is not None:
        primeira_coluna_denom = dados[ind_total].filter((pl.col('Year')==year_ultimo)&(pl.col('Month')==month_ultimo))
        primeira_coluna_numer = dados[ind_total].filter((pl.col('Year')==year_antepenultimo)&(pl.col('Month')==month_penultimo))   
        ultima_data = dados[ind_total].sort(by=["Data"])[-1]["Data"][0]
        perc = primeira_coluna_denom["New_Users"]/primeira_coluna_numer["New_Users"] - 1.0
        ws.cell(row=linha, column=coluna, value=perc[0]).style=normalgrayperc
        soma_denom = sum([dados[ind_total].filter((pl.col('Year')==lista_anos[i])&(pl.col('Month')==lista_meses[i]))[ind_col] for i in range(len(lista_meses))])
        perc = primeira_coluna_denom["New_Users"]/(soma_denom / 4.0) - 1.0
        ws.cell(row=linha, column=coluna+1, value=perc[0]).style=normalgrayperc
        total_denom = dados[ind_total].filter(pl.col('Data')<=ultima_data)["New_Users"].sum()
        perc = primeira_coluna_denom["New_Users"]/total_denom
        ws.cell(row=linha, column=coluna+2, value=perc[0]).style=normalgrayperc
    for i, item in enumerate(lista_segm):
        primeira_coluna_denom = dados[indice].filter((pl.col('Year')==year_ultimo)&(pl.col('Month')==month_ultimo))[item]
        primeira_coluna_numer = dados[indice].filter((pl.col('Year')==year_penultimo)&(pl.col('Month')==month_penultimo))[item]
        perc = primeira_coluna_denom/primeira_coluna_numer - 1.0
        ws.cell(row=linha+i+1, column=coluna, value=perc[0]).style=(normalgrayperc if i+linha < listagem_maximo else normalgrayunderperc)
        soma_denom = sum([dados[indice].filter((pl.col('Year')==lista_anos[j])&(pl.col('Month')==lista_meses[j]))[item] for j in range(len(lista_meses))])
        perc = float(primeira_coluna_denom[0])/(float(soma_denom[0])/4.0) - 1.0 if soma_denom[0] != 0.0 else ''
        ws.cell(row=linha+i+1, column=coluna+1, value=perc).style=(normalgrayperc if i+linha < listagem_maximo else normalgrayunderperc)
        total_denom = dados[indice].filter(pl.col('Data')<ultima_data)[item].sum()
        perc = primeira_coluna_denom[0]/total_denom if total_denom != 0.0 else ''
        ws.cell(row=linha+i+1, column=coluna+2, value=perc).style=(normalgrayperc if i+linha < listagem_maximo else normalgrayunderperc)
        
def final_values(ultimo: datetime, linha: int,coluna: int, ws: worksheet, dados: dict[str, pl.DataFrame], ident: str, ident_tot: str, lista_segm: list[str])->None:
    ws.cell(row=linha, column=coluna, value = dados[ident_tot].sort(by=["Data"])["New_Users"][0]).style=normalgray
    for i, seg in enumerate(lista_segm):
        ws.cell(row=linha+i+1, column=coluna, value = dados[ident][seg].to_list()[0]).style=normalgray
