from openpyxl import Workbook
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import NamedStyle, Side
wb = Workbook()
ws = wb.active

ws['A3'] = (datetime.now() - datetime(1900, 1, 1)).days
ws['A4'] = ((datetime.now()-timedelta(days=1)) - datetime(1900, 1, 1)).days
ws['A5'] = ((datetime.now()-timedelta(days=2)) - datetime(1900, 1, 1)).days

ws['A3'].number_format='yyyy-mm-dd'
ws['A4'].number_format='yyyy-mm-dd'
ws['A5'].number_format='yyyy-mm-dd'

highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True, size=9)
bd = Side(style='thick', color="000000")
highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
highlight.number_format = 'yyyy-mm-dd'
wb.add_named_style(highlight)

normal = NamedStyle(name="normal")
normal.font = Font(bold=False, size=8)
normal.number_format = 'yyyy-mm-dd'
wb.add_named_style(normal)

ws['A3'].style = highlight
ws['A4'].style = highlight
ws['A5'].style = normal

wb.save("sample.xlsx")